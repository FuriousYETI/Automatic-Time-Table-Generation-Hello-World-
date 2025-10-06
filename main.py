import pandas as pd
import json
import random
import colorsys
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

DATA_DIR = Path("data")
TIME_SLOTS_FILE = DATA_DIR / "time_slots.json"
COURSES_FILE = DATA_DIR / "courses.csv"
ROOMS_FILE = DATA_DIR / "rooms.csv"
DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
EXCLUDED_SLOTS = ["07:30-09:00", "13:15-14:00", "17:30-18:30"]
MAX_ATTEMPTS = 12

def load_time_slots(path):
    with open(path) as f:
        slots = json.load(f)["time_slots"]
    keys = [f"{s['start'].strip()}-{s['end'].strip()}" for s in slots]
    return keys

SLOT_KEYS = load_time_slots(TIME_SLOTS_FILE)

def slot_duration(slot):
    start, end = slot.split("-")
    h1, m1 = map(int, start.split(":"))
    h2, m2 = map(int, end.split(":"))
    return (h2 + m2 / 60) - (h1 + m1 / 60)

SLOT_DURATIONS = {s: slot_duration(s) for s in SLOT_KEYS}

def generate_palette(n, light_range=(0.45,0.62), sat_range=(0.68,0.9)):
    palette = []
    for i in range(n):
        hue = i / max(1, n)
        light = random.uniform(*light_range)
        sat = random.uniform(*sat_range)
        r, g, b = colorsys.hls_to_rgb(hue, light, sat)
        palette.append(f"{int(r*255):02X}{int(g*255):02X}{int(b*255):02X}")
    random.shuffle(palette)
    return palette

class Course:
    def __init__(self, record):
        self.record = record
        self.code = str(record.get("Course_Code","")).strip()
        self.faculty = str(record.get("Faculty","")).strip()
        self.sem_half = str(record.get("Semester_Half","")).strip()
        self.is_elective_flag = str(record.get("Elective",0)).strip() == "1"
        try:
            L,T,P,S,C = map(int, [x.strip() for x in record.get("L-T-P-S-C","0-0-0-0-0").split("-")])
        except:
            L,T,P,S,C = 0,0,0,0,0
        self.L = L
        self.T = T
        self.P = P

class RoomPool:
    def __init__(self, rooms_df):
        self.classrooms = rooms_df[rooms_df["Type"].str.lower()=="classroom"]["Room_ID"].tolist()
        self.labs = rooms_df[rooms_df["Type"].str.lower()=="lab"]["Room_ID"].tolist()
    def pick(self, session_type):
        if session_type == "P":
            return random.choice(self.labs) if self.labs else None
        return random.choice(self.classrooms) if self.classrooms else None

class TimetableGenerator:
    def __init__(self, courses, rooms, slot_keys=SLOT_KEYS, days=DAYS, excluded=EXCLUDED_SLOTS):
        self.courses = [Course(c) for c in courses]
        self.room_pool = RoomPool(rooms)
        self.slot_keys = slot_keys
        self.days = days
        self.excluded = set(excluded)
        self.course_room_map = {}
        self.palette = generate_palette(64)
    def _new_empty(self):
        return pd.DataFrame("", index=self.days, columns=self.slot_keys)
    def _get_free_blocks(self, table, day):
        free_blocks = []
        block = []
        for s in self.slot_keys:
            if table.at[day, s] == "" and s not in self.excluded:
                block.append(s)
            else:
                if block:
                    free_blocks.append(block)
                    block = []
        if block:
            free_blocks.append(block)
        return free_blocks
    def _assign_room(self, code, session_type):
        if code in self.course_room_map:
            return self.course_room_map[code]
        room = self.room_pool.pick(session_type)
        if room is None:
            return None
        self.course_room_map[code] = room
        return room
    def _allocate(self, table, busy, day, faculty, code, dur_hours, session_type="L", is_elective=False):
        blocks = self._get_free_blocks(table, day)
        for block in blocks:
            total = sum(SLOT_DURATIONS[s] for s in block)
            if total >= dur_hours:
                selected = []
                acc = 0
                for s in block:
                    selected.append(s)
                    acc += SLOT_DURATIONS[s]
                    if acc >= dur_hours:
                        break
                room = None
                if not is_elective:
                    room = self._assign_room(code, session_type)
                    if room is None:
                        return False
                for s in selected:
                    if session_type == "L":
                        table.at[day, s] = f"{code} ({room})" if not is_elective else code
                    elif session_type == "T":
                        table.at[day, s] = f"{code}T ({room})" if not is_elective else f"{code}T"
                    else:
                        table.at[day, s] = f"{code} (Lab-{room})" if not is_elective else code
                if faculty:
                    busy[day].append(faculty)
                return True
        return False
    def _place_course(self, table, course, busy):
        faculty = course.faculty
        code = course.code if course.code else "Unknown"
        is_elective = course.is_elective_flag or code.lower()=="elective"
        try:
            lecture = course.L
            tut = course.T
            prac = course.P
        except:
            return
        attempts = 0
        while lecture > 0 and attempts < MAX_ATTEMPTS:
            attempts += 1
            for day in self.days:
                if lecture <= 0 or (faculty and faculty in busy[day]):
                    continue
                alloc = min(1.5, lecture)
                if self._allocate(table, busy, day, faculty, code, alloc, "L", is_elective):
                    lecture -= alloc
                    break
        attempts = 0
        while tut > 0 and attempts < MAX_ATTEMPTS:
            attempts += 1
            for day in self.days:
                if tut <= 0 or (faculty and faculty in busy[day]):
                    continue
                if self._allocate(table, busy, day, faculty, code, 1, "T", is_elective):
                    tut -= 1
                    break
        attempts = 0
        while prac > 0 and attempts < MAX_ATTEMPTS:
            attempts += 1
            for day in self.days:
                if prac <= 0 or (faculty and faculty in busy[day]):
                    continue
                alloc = min(2, prac)
                if self._allocate(table, busy, day, faculty, code, alloc, "P", is_elective):
                    prac -= alloc
                    break
    def _finalize_excluded(self, table):
        for d in self.days:
            for s in self.excluded:
                if s in table.columns:
                    table.at[d, s] = ""
    def generate(self, sem_half="1"):
        records = [c for c in self.courses if (str(c.sem_half).strip() in [sem_half,"0"])]
        electives = [c for c in records if c.is_elective_flag]
        non_electives = [c for c in records if not c.is_elective_flag]
        if electives:
            chosen = random.choice(electives)
            wrapper = Course({"Course_Code":"Elective","Faculty":chosen.faculty,"L-T-P-S-C":f"{chosen.L}-{chosen.T}-{chosen.P}-0-0","Semester_Half":sem_half,"Elective":1})
            non_electives.append(wrapper)
        table = self._new_empty()
        busy = {d:[] for d in self.days}
        for c in non_electives:
            self._place_course(table, c, busy)
        self._finalize_excluded(table)
        return table, dict(self.course_room_map)

class ExcelStyler:
    def __init__(self, palette=None):
        self.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        self.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.font_dark = Font(bold=True, color="000000")
        self.font_light = Font(bold=True, color="FFFFFF")
        self.palette = palette or generate_palette(64)
    def _choose_text_color(self, hexcolor):
        r = int(hexcolor[0:2],16); g = int(hexcolor[2:4],16); b = int(hexcolor[4:6],16)
        luminance = (0.299*r + 0.587*g + 0.114*b)/255
        return self.font_dark if luminance > 0.55 else self.font_light
    def _apply_fill(self, cell, hexcolor):
        cell.fill = PatternFill(start_color=hexcolor, end_color=hexcolor, fill_type="solid")
        cell.alignment = self.alignment
        cell.font = self._choose_text_color(hexcolor)
        cell.border = self.border
    def style_and_save(self, df, filepath, course_room_map, title="Timetable"):
        df.to_excel(filepath, index=True)
        wb = load_workbook(filepath)
        ws = wb.active
        course_colors = {}
        available = self.palette.copy()
        random.shuffle(available)
        max_row = ws.max_row
        max_col = ws.max_column
        for r in range(2, max_row+1):
            c = 2
            while c <= max_col:
                val = ws.cell(row=r, column=c).value
                if not val:
                    c += 1
                    continue
                merge_cols = [c]
                session_dur = 1.5
                sval = str(val)
                if "(" in sval:
                    if "Lab" in sval:
                        session_dur = 2
                    elif "T" in sval and sval.strip().endswith("T"):
                        session_dur = 1
                    else:
                        session_dur = 1.5
                base_dur = SLOT_DURATIONS[SLOT_KEYS[c-2]]
                total = base_dur
                nc = c+1
                while nc <= max_col:
                    nv = ws.cell(row=r, column=nc).value
                    if nv == val:
                        total += SLOT_DURATIONS[SLOT_KEYS[nc-2]]
                        merge_cols.append(nc)
                        if total >= session_dur:
                            break
                        nc += 1
                    else:
                        break
                if len(merge_cols) > 1:
                    ws.merge_cells(start_row=r, start_column=merge_cols[0], end_row=r, end_column=merge_cols[-1])
                main_cell = ws.cell(row=r, column=merge_cols[0])
                main_cell.alignment = self.alignment
                course_key = str(val).split()[0].replace("T","")
                if course_key not in course_colors:
                    course_colors[course_key] = available.pop() if available else "%02X%02X%02X" % (random.randint(60,200),random.randint(60,200),random.randint(60,200))
                hexcol = course_colors[course_key]
                self._apply_fill(main_cell, hexcol)
                for cc in merge_cols:
                    ws.cell(row=r, column=cc).border = self.border
                c = merge_cols[-1] + 1
        legend_ws = wb.create_sheet("Legend")
        legend_ws.append(["Course", "Room", "Color (HEX)"])
        for course_key, hexcol in sorted(course_colors.items()):
            room = course_room_map.get(course_key, "")
            legend_ws.append([course_key, room, hexcol])
        tbl = Table(displayName="LegendTable", ref=f"A1:C{legend_ws.max_row}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        legend_ws.add_table(tbl)
        for col in range(1,4):
            legend_ws.column_dimensions[get_column_letter(col)].width = 20
            for r in range(1, legend_ws.max_row+1):
                cell = legend_ws.cell(row=r, column=col)
                cell.alignment = self.alignment
        wb.save(filepath)

def main():
    courses_df = pd.read_csv(COURSES_FILE)
    rooms_df = pd.read_csv(ROOMS_FILE)
    gen = TimetableGenerator(courses_df.to_dict(orient="records"), rooms_df)
    styler = ExcelStyler(palette=gen.palette)
    first_half = [c for c in courses_df.to_dict(orient="records") if str(c.get("Semester_Half","")).strip() in ["1","0"]]
    second_half = [c for c in courses_df.to_dict(orient="records") if str(c.get("Semester_Half","")).strip() in ["2","0"]]
    t1, map1 = gen.generate(sem_half="1")
    t2, map2 = gen.generate(sem_half="2")
    out1 = "TT_first_halfsem.xlsx"
    out2 = "TT_second_halfsem.xlsx"
    styler.style_and_save(t1, out1, map1, title="First Half Timetable")
    styler.style_and_save(t2, out2, map2, title="Second Half Timetable")
    print(f"Saved: {out1}, {out2}")

if __name__ == "__main__":
    main()
