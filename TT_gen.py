import pandas as pd
import json
import os
import random
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# Use a fresh seed each run to get a new timetable
random.seed()
print("Generating Timetable...")
days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
excluded = ["07:30-09:00", "10:30-10:45", "13:15-14:00", "17:30-18:30"]
# Tracks which course is using C004 in each slot (across all years/branches)
c004_occupancy = {d: {} for d in days}   # day -> {slot -> course_code}
# never allow any placement in these slots (hard ban)
ABSOLUTELY_FORBIDDEN_SLOTS = {"07:30-09:00", "17:30-18:30"}
MINOR_SLOTS = {"07:30-09:00", "17:30-18:30"}
BREAK_SLOTS = {"10:30-10:45", "13:15-14:00"}
# Never schedule anything in these slots (even when ex=True)
HARD_FORBIDDEN_SLOTS = ABSOLUTELY_FORBIDDEN_SLOTS | BREAK_SLOTS


colors = [
    "FFB3BA","BAE1FF","BAFFC9","FFFFBA","FFD8BA","E3BAFF","D0BAFF","FFCBA4",
    "C7FFD8","B8E1FF","F7FFBA","FFDFBA","E9BAFF","BAFFD9","FFE1BA","BAFFF2",
    "D1FFBA","B2D8F7","F2C2FF","C2FFD8","FFB8E1","D8FFB8","FFE3BA","BAE7FF",
    "E8BAFF","BAFFD6","FFF2BA","DAD7FF","BFFFE1","FFDAB8","E2FFBA","BAF7FF"
]

thin = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))

with open("data/time_slots.json") as f:
    slots = json.load(f)["time_slots"]

def normalize_time(t):
    h, m = t.split(":")
    return f"{int(h):02d}:{int(m):02d}"

def t2m(t):
    h, m = map(int, t.split(":"))
    return h*60 + m

def shorten_faculty_name(name):
    if name is None:
        return ""
    sname = str(name).strip()
    if not sname:
        return sname
    parts = [p.strip() for p in sname.split("/")]
    return " / ".join(_shorten_faculty_single(p) for p in parts if p)

def _shorten_faculty_single(name):
    n = re.sub(r"\s+", " ", str(name)).strip()
    if not n:
        return n
    titles = ("Dr.", "Prof.", "Mr.", "Ms.", "Mrs.")
    for t in titles:
        if n.startswith(t) and len(n) > len(t) and n[len(t)] != " ":
            n = t + " " + n[len(t):]
            break
    tokens = n.split()
    title = ""
    if tokens and tokens[0] in titles:
        title = tokens[0]
        tokens = tokens[1:]
    if not tokens:
        return title or n
    first = tokens[0]
    rest = tokens[1:]
    initials = []
    for tok in rest:
        m = re.search(r"[A-Za-z]", tok)
        if not m:
            continue
        initials.append(m.group(0).upper())
    if title:
        return f"{title} {first}" + (f" {' '.join(initials)}" if initials else "")
    return f"{first}" + (f" {' '.join(initials)}" if initials else "")

def split_faculty_names(name):
    if name is None:
        return []
    parts = [p.strip() for p in str(name).split("/") if p.strip()]
    return parts

def build_course_index():
    """Map Course_Code -> list of (Course_Title, Departments, Faculty)."""
    idx = {}
    all_courses = coursesCSE + coursesECE + coursesDSAI + (coursesVII or [])
    for c in all_courses:
        code = s(c.get("Course_Code",""))
        title = s(c.get("Course_Title",""))
        dept = s(c.get("Departments",""))
        fac = s(c.get("Faculty",""))
        if not code:
            continue
        idx.setdefault(code, []).append((title, dept, fac))
    return idx

def build_room_map_from_tt(tt):
    """
    Build a map: Course_Code -> set of rooms found in timetable cells.
    """
    room_map = {}
    for d in days:
        for s_ in slot_keys:
            val = tt.at[d, s_]
            if not isinstance(val, str) or val.strip() == "":
                continue
            code = extract_course_code(val)
            if not code:
                continue
            m = re.search(r"\(([^)]+)\)", val)
            if not m:
                continue
            room = m.group(1).strip()
            if not room:
                continue
            room_map.setdefault(code, set()).add(room)
    return room_map

def build_course_faculty_map():
    """Map Course_Code -> list of faculty names (split on '/')."""
    fmap = {}
    all_courses = coursesCSE + coursesECE + coursesDSAI + (coursesVII or [])
    for c in all_courses:
        code = s(c.get("Course_Code",""))
        fac = s(c.get("Faculty",""))
        if not code or not fac:
            continue
        fmap[code] = split_faculty_names(fac)
    return fmap

def _parse_blocks_from_ws(ws):
    """
    Return list of blocks: dict with keys:
    label, half (1 or 2), header_row, slots(list), day_rows(list of (row_idx, day))
    """
    blocks = []
    max_row = ws.max_row
    for r in range(1, max_row + 1):
        val = ws.cell(r, 1).value
        if not isinstance(val, str):
            continue
        if "First Half" in val or "Second Half" in val:
            label = val.strip()
            half = 1 if "First Half" in label else 2
            # find header row with "Day"
            hr = r + 1
            while hr <= max_row and str(ws.cell(hr, 1).value).strip() != "Day":
                hr += 1
            if hr > max_row:
                continue
            # slot headers from column 2 until empty
            slots = []
            c = 2
            while True:
                v = ws.cell(hr, c).value
                if v is None or str(v).strip() == "":
                    break
                slots.append(str(v).strip())
                c += 1
            # day rows until blank day
            day_rows = []
            rr = hr + 1
            while rr <= max_row:
                day = ws.cell(rr, 1).value
                if not isinstance(day, str) or day.strip() == "":
                    break
                day_rows.append((rr, day.strip()))
                rr += 1
            blocks.append({
                "label": label,
                "half": half,
                "header_row": hr,
                "slots": slots,
                "day_rows": day_rows
            })
    return blocks

def repair_faculty_clashes(wb, faculty_tt, course_faculty_map):
    """
    Deterministic repair: for each half, if same faculty teaches different
    courses in same day+slot across blocks, move later-sorted entries to any
    free contiguous slot block in the same timetable block (prefer same day).
    """
    moved = 0
    entries = []

    # Build entries and faculty usage map
    from collections import defaultdict
    fac_usage = defaultdict(list)  # (half, day, slot, faculty) -> list of entries
    for ws in wb.worksheets:
        blocks = _parse_blocks_from_ws(ws)
        for blk in blocks:
            slots = blk["slots"]
            slot_idx = {s:i for i,s in enumerate(slots)}
            for row_idx, day in blk["day_rows"]:
                for s in slots:
                    col = 2 + slot_idx[s]
                    val = ws.cell(row_idx, col).value
                    if not isinstance(val, str) or val.strip() == "":
                        continue
                    code = extract_course_code(val)
                    if not code:
                        continue
                    facs = course_faculty_map.get(code, [])
                    e = {
                        "half": blk["half"],
                        "day": day,
                        "slot": s,
                        "faculty_list": facs,
                        "ws": ws,
                        "block": blk,
                        "row": row_idx,
                        "col": col,
                        "value": val,
                        "code": code
                    }
                    entries.append(e)
                    for fac in facs:
                        fac_usage[(blk["half"], day, s, fac)].append(e)

    # group by clash key
    groups = defaultdict(list)
    for e in entries:
        for fac in e["faculty_list"]:
            groups[(e["half"], e["day"], e["slot"], fac)].append(e)

    def is_slot_free_for_fac(half, day, slot, fac_list):
        for fac in fac_list:
            if fac_usage.get((half, day, slot, fac)):
                return False
        return True

    def is_merged(ws, row, col):
        return ws.cell(row, col).coordinate in ws.merged_cells

    def find_contiguous_block(ws, row, slots, slot_idx, code):
        # find contiguous block of same course code in the row
        idxs = []
        for s in slots:
            col = 2 + slot_idx[s]
            if is_merged(ws, row, col):
                return None
            val = ws.cell(row, col).value
            if isinstance(val, str) and extract_course_code(val) == code:
                idxs.append(slot_idx[s])
        if not idxs:
            return None
        idxs.sort()
        # return the contiguous run containing the first idx
        start = idxs[0]
        end = idxs[0]
        for i in idxs[1:]:
            if i == end + 1:
                end = i
            else:
                break
        return list(range(start, end + 1))

    for key, items in groups.items():
        # only if different courses
        codes = {i["code"] for i in items}
        if len(codes) <= 1:
            continue
        # deterministic order: keep first, move others
        items.sort(key=lambda x: (x["ws"].title, x["block"]["label"], x["code"]))
        for it in items[1:]:
            blk = it["block"]
            ws = it["ws"]
            slots = blk["slots"]
            slot_idx = {s:i for i,s in enumerate(slots)}
            # contiguous block to move
            run = find_contiguous_block(ws, it["row"], slots, slot_idx, it["code"])
            if not run:
                continue
            run_len = len(run)
            half = blk["half"]
            fac_list = it["faculty_list"]

            # helper to test a candidate start index on a given row
            def can_place(row_idx, start_i):
                target_slots = slots[start_i:start_i+run_len]
                # slot availability + forbidden check
                for s in target_slots:
                    if s in FORBIDDEN_SLOTS:
                        return False
                    col = 2 + slot_idx[s]
                    if is_merged(ws, row_idx, col):
                        return False
                    if ws.cell(row_idx, col).value not in (None, ""):
                        return False
                    if not is_slot_free_for_fac(half, day_name, s, fac_list):
                        return False
                return True

            placed = False
            # try same day first
            day_name = it["day"]
            for start_i in range(0, len(slots) - run_len + 1):
                if can_place(it["row"], start_i):
                    target_slots = slots[start_i:start_i+run_len]
                    # move cells
                    for idx in run:
                        s = slots[idx]
                        col = 2 + slot_idx[s]
                        if is_merged(ws, it["row"], col):
                            continue
                        ws.cell(it["row"], col).value = ""
                        for fac in fac_list:
                            fac_usage[(half, day_name, s, fac)] = [
                                e for e in fac_usage[(half, day_name, s, fac)] if e is not it
                            ]
                            if half in faculty_tt and fac in faculty_tt[half] and day_name in faculty_tt[half][fac]:
                                faculty_tt[half][fac][day_name].pop(s, None)
                    for s in target_slots:
                        col = 2 + slot_idx[s]
                        if is_merged(ws, it["row"], col):
                            continue
                        ws.cell(it["row"], col).value = it["value"]
                        for fac in fac_list:
                            fac_usage[(half, day_name, s, fac)].append(it)
                            if half in faculty_tt:
                                faculty_tt[half].setdefault(fac, {}).setdefault(day_name, {})[s] = it["value"]
                    moved += 1
                    placed = True
                    break
            if placed:
                continue

            # try other days
            for row_idx, day_name in blk["day_rows"]:
                if day_name == it["day"]:
                    continue
                for start_i in range(0, len(slots) - run_len + 1):
                    if can_place(row_idx, start_i):
                        target_slots = slots[start_i:start_i+run_len]
                        # clear old
                        for idx in run:
                            s = slots[idx]
                            col = 2 + slot_idx[s]
                            if is_merged(ws, it["row"], col):
                                continue
                            ws.cell(it["row"], col).value = ""
                            for fac in fac_list:
                                fac_usage[(half, it["day"], s, fac)] = [
                                    e for e in fac_usage[(half, it["day"], s, fac)] if e is not it
                                ]
                                if half in faculty_tt and fac in faculty_tt[half] and it["day"] in faculty_tt[half][fac]:
                                    faculty_tt[half][fac][it["day"]].pop(s, None)
                        # place new
                        for s in target_slots:
                            col = 2 + slot_idx[s]
                            if is_merged(ws, row_idx, col):
                                continue
                            ws.cell(row_idx, col).value = it["value"]
                            for fac in fac_list:
                                fac_usage[(half, day_name, s, fac)].append(it)
                                if half in faculty_tt:
                                    faculty_tt[half].setdefault(fac, {}).setdefault(day_name, {})[s] = it["value"]
                        moved += 1
                        placed = True
                        break
                if placed:
                    break

    return moved

def course_key(c):
    return (
        s(c.get("Departments","")),
        s(c.get("Semester","")),
        s(c.get("Section","")),
        s(c.get("Course_Code","")),
        s(c.get("Course_Title","")),
        s(c.get("Faculty","")),
        s(c.get("Semester_Half","")),
        s(c.get("L-T-P-S-C","")),
    )

def full_sem_key(c, year_tag):
    return (
        year_tag,
        s(c.get("Departments","")),
        s(c.get("Section","")),
        s(c.get("Course_Code","")),
    )

def _classify_slot_val(code, val):
    v = str(val)
    code_u = code.strip().upper()
    v_u = v.strip().upper()
    if "LAB" in v_u:
        return "P"
    if "TUT" in v_u or v_u.startswith(code_u + "T"):
        return "T"
    return "L"

def build_full_sem_sync_from_tt(tt, courses, year_tag, full_sem_sync):
    for c in courses:
        if s(c.get("Semester_Half","")) != "0":
            continue
        code = s(c.get("Course_Code",""))
        if not code:
            continue
        fs_key = full_sem_key(c, year_tag)
        by_typ = {"L": [], "T": [], "P": []}
        for d in days:
            cur_slots = []
            cur_typ = None
            for s_ in slot_keys:
                val = tt.at[d, s_]
                if isinstance(val, str) and val.strip().upper().startswith(code.upper()):
                    typ = _classify_slot_val(code, val)
                    if cur_typ is None:
                        cur_typ = typ
                        cur_slots = [s_]
                    elif typ == cur_typ:
                        cur_slots.append(s_)
                    else:
                        if cur_slots:
                            by_typ[cur_typ].append((d, cur_slots))
                        cur_typ = typ
                        cur_slots = [s_]
                else:
                    if cur_slots:
                        by_typ[cur_typ].append((d, cur_slots))
                        cur_slots = []
                        cur_typ = None
            if cur_slots:
                by_typ[cur_typ].append((d, cur_slots))
        full_sem_sync[fs_key] = by_typ

def collect_code_slot_blocks(tt, code):
    blocks = []
    if not code:
        return blocks
    code_u = code.strip().upper()
    for d in days:
        cur_slots = []
        for s_ in slot_keys:
            val = tt.at[d, s_]
            if isinstance(val, str) and val.strip().upper().startswith(code_u):
                cur_slots.append(s_)
            else:
                if cur_slots:
                    blocks.append((d, cur_slots))
                    cur_slots = []
        if cur_slots:
            blocks.append((d, cur_slots))
    return blocks

def normalize_elective_basket(year_tag, basket_key):
    """
    Normalize elective basket keys for scheduling rules.
    Keep basket identifiers distinct so LTPSC is enforced per basket.
    """
    return basket_key

def apply_basket_rooms_to_tt(tt, year_tag):
    if year_tag is None:
        return
    for d in days:
        for s_ in slot_keys:
            val = tt.at[d, s_]
            if not isinstance(val, str) or not val:
                continue
            if "(" in val:
                continue
            code_u = val.strip().upper()
            if not code_u.startswith("ELECTIVE"):
                continue
            basket_num = _basket_code_parts(val)
            if not basket_num:
                continue
            rooms = get_basket_room_list(year_tag, basket_num)
            if rooms:
                tt.at[d, s_] = f"{val} ({', '.join(rooms)})"

def collect_unscheduled(courses, placed_list, group_label, year_tag=None, elective_sync=None):
    placed_keys = set(course_key(c) for c in placed_list if isinstance(c, dict))
    uns = []
    for c in courses:
        # If elective belongs to a basket that is scheduled, treat as scheduled
        if s(c.get("Elective","")) == "1":
            basket = s(c.get("ElectiveBasket","0"))
            basket = normalize_elective_basket(year_tag, basket)
            if basket and basket != "0" and elective_sync is not None:
                sync_identifier = f"Y{year_tag}_B{basket}" if year_tag is not None else f"B{basket}"
                if sync_identifier in elective_sync:
                    continue
        if course_key(c) not in placed_keys:
            uns.append({
                "Group": group_label,
                "Department": s(c.get("Departments","")),
                "Semester": s(c.get("Semester","")),
                "Section": s(c.get("Section","")),
                "Course_Code": s(c.get("Course_Code","")),
                "Course_Title": s(c.get("Course_Title","")),
                "Faculty": shorten_faculty_name(c.get("Faculty","")),
                "L-T-P-S-C": s(c.get("L-T-P-S-C","")),
                "Elective": s(c.get("Elective","")),
                "ElectiveBasket": s(c.get("ElectiveBasket","")),
                "Semester_Half": s(c.get("Semester_Half","")),
            })
    return uns

slots_norm = [
    {
        "key": f"{normalize_time(s['start'])}-{normalize_time(s['end'])}",
        "start": normalize_time(s['start']),
        "end": normalize_time(s['end']),
        "dur": (t2m(s["end"]) - t2m(s["start"])) / 60.0
    }
    for s in slots
]
slots_norm.sort(key=lambda x: t2m(x["start"]))
slot_keys = [s["key"] for s in slots_norm]
slot_dur = {s["key"]: s["dur"] for s in slots_norm}
FORBIDDEN_SLOTS = set(excluded) | ABSOLUTELY_FORBIDDEN_SLOTS

#############################################
# NEW EXCEL INPUT LOADER (YOUR FORMAT)
#############################################

REQUIRED_COLUMNS = [
    "Semester","Section",
    "Course code","Course name",
    "L","T","P","S","C","Faculty",
    "Combined","Elective","total_students","ElectiveBasket","Semester_Half"
]

def load_and_validate(file):
    df = pd.read_csv(file)
    # check columns
    # if Section column missing -> auto create
    if "Section" not in df.columns:
        df["Section"] = "ALL"

    # normalize schedule column name if needed
    if "Schedule" not in df.columns and "Schedule(all columns yes)" in df.columns:
        df.rename(columns={"Schedule(all columns yes)":"Schedule"}, inplace=True)

    # if Departments missing, infer from filename
    if "Departments" not in df.columns:
        fname = str(file).upper()
        if "CSE" in fname:
            df["Departments"] = "CSE"
        elif "ECE" in fname:
            df["Departments"] = "ECE"
        elif "DSAI" in fname:
            df["Departments"] = "DSAI"
        else:
            df["Departments"] = ""

    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise Exception(f"{file} missing columns: {missing}")

    df = df.fillna("")
    df["Section"] = df["Section"].astype(str).str.strip()

    # numeric columns
    for col in ["L","T","P","S","C","Semester","total_students","Elective","ElectiveBasket","Semester_Half"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # build L-T-P-S-C
    df["L-T-P-S-C"] = (
        df["L"].astype(int).astype(str)+"-"+
        df["T"].astype(int).astype(str)+"-"+
        df["P"].astype(int).astype(str)+"-"+
        df["S"].astype(int).astype(str)+"-"+
        df["C"].astype(int).astype(str)
    )

    # rename for engine
    df.rename(columns={
        "Course code":"Course_Code",
        "Course name":"Course_Title"
    }, inplace=True)

    # defaults needed by engine (only if not already present)
    if "Elective" not in df.columns:
        df["Elective"] = 0
    if "Is_Combined" not in df.columns:
        df["Is_Combined"] = 0
    if "ElectiveBasket" not in df.columns:
        df["ElectiveBasket"] = 0
    # If Combined column is provided, map it to Is_Combined
    if "Combined" in df.columns:
        df["Is_Combined"] = pd.to_numeric(df["Combined"], errors="coerce").fillna(0)
    df["Is_Combined"] = pd.to_numeric(df["Is_Combined"], errors="coerce").fillna(0).astype(int)
    # If Semester_Half is not provided, derive it from C:
    # C <= 2 => half semester (1), C > 2 => full semester (0)
    if "Semester_Half" not in df.columns:
        df["Semester_Half"] = df["C"].apply(lambda x: 1 if float(x) <= 2 else 0)

    return df.to_dict(orient="records")

def load_and_validate_sem7(file):
    df = pd.read_csv(file)

    # Normalize column names from Course7.xlsx/CSV format
    rename_map = {
        "Course_Code": "Course code",
        "Course_Title": "Course name",
        "Students": "total_students",
        "ElectiveBas": "ElectiveBasket"
    }
    for k, v in rename_map.items():
        if k in df.columns and v not in df.columns:
            df.rename(columns={k: v}, inplace=True)

    # Required defaults for engine
    if "Semester" not in df.columns:
        df["Semester"] = 7
    if "Section" not in df.columns:
        df["Section"] = "ALL"
    if "Departments" not in df.columns:
        df["Departments"] = "COMMON"
    if "Combined" not in df.columns:
        df["Combined"] = 0
    if "Elective" not in df.columns:
        df["Elective"] = 1
    if "ElectiveBasket" not in df.columns:
        df["ElectiveBasket"] = 0
    if "total_students" not in df.columns:
        df["total_students"] = 0
    if "Semester_Half" not in df.columns:
        df["Semester_Half"] = 0

    missing = [c for c in ["Course code","Course name","L","T","P","S","C","Faculty"] if c not in df.columns]
    if missing:
        raise Exception(f"{file} missing columns: {missing}")

    df = df.fillna("")
    df["Section"] = df["Section"].astype(str).str.strip()

    # numeric columns
    for col in ["L","T","P","S","C","Semester","total_students","Elective","ElectiveBasket","Semester_Half","Combined"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # build L-T-P-S-C
    df["L-T-P-S-C"] = (
        df["L"].astype(int).astype(str)+"-"+
        df["T"].astype(int).astype(str)+"-"+
        df["P"].astype(int).astype(str)+"-"+
        df["S"].astype(int).astype(str)+"-"+
        df["C"].astype(int).astype(str)
    )

    # rename for engine
    df.rename(columns={
        "Course code":"Course_Code",
        "Course name":"Course_Title"
    }, inplace=True)

    # defaults needed by engine
    if "Is_Combined" not in df.columns:
        df["Is_Combined"] = 0
    # If Combined column is provided, map it to Is_Combined
    if "Combined" in df.columns:
        df["Is_Combined"] = pd.to_numeric(df["Combined"], errors="coerce").fillna(0)
    df["Is_Combined"] = pd.to_numeric(df["Is_Combined"], errors="coerce").fillna(0).astype(int)

    return df.to_dict(orient="records")


# LOAD COURSE DETAILS
coursesCSE  = load_and_validate("data/CSE_courses.csv")
coursesECE  = load_and_validate("data/ECE_courses.csv")
coursesDSAI = load_and_validate("data/DSAI_courses.csv")


##########################################
#              SPLITING DATA             #
##########################################
def filter_courses(data, dept, sem, section=None):
    res = []

    for c in data:
        # department match
        if str(c.get("Departments","")).strip().upper() != dept.upper():
            continue

        # semester match
        try:
            s = int(float(str(c.get("Semester",0)).strip()))
        except:
            continue
        if s != sem:
            continue

        # SECTION FILTER ONLY FOR CSE
        if dept.upper() == "CSE" and section in ("A","B"):
            sec = str(c.get("Section","")).strip().upper()
            if sec != section:
                continue

        res.append(c)

    return res

# CSE
coursesCSEA_I = filter_courses(coursesCSE,"CSE",1,"A")
coursesCSEB_I = filter_courses(coursesCSE,"CSE",1,"B")

# sem3
coursesCSEA_III = filter_courses(coursesCSE,"CSE",3,"A")
coursesCSEB_III = filter_courses(coursesCSE,"CSE",3,"B")

# sem5
coursesCSEA_V = filter_courses(coursesCSE,"CSE",5,"A")
# CSEB sem5
coursesCSEB_V = filter_courses(coursesCSE,"CSE",5,"B")
# sem7 (if exists in CSE)
coursesCSEA_VII = filter_courses(coursesCSE,"CSE",7,"A")
# DSAI
coursesDSAI_I = filter_courses(coursesDSAI,"DSAI",1)
coursesDSAI_III = filter_courses(coursesDSAI,"DSAI",3)
coursesDSAI_V   = filter_courses(coursesDSAI,"DSAI",5)

# ECE
coursesECE_I = filter_courses(coursesECE,"ECE",1)
coursesECE_III = filter_courses(coursesECE,"ECE",3)
coursesECE_V   = filter_courses(coursesECE,"ECE",5)

# dummy for engine
coursesAI = coursesCSEA_I
coursesBI = coursesCSEB_I

coursesA  = coursesCSEA_III
coursesB  = coursesCSEB_III

coursesV  = coursesCSEA_V
coursesVII = []
if os.path.exists("data/Course7.csv"):
    coursesVII = load_and_validate_sem7("data/Course7.csv")
elif 'coursesCSEA_VII' in globals():
    coursesVII = coursesCSEA_VII

rooms = pd.read_csv("data/rooms.csv")
rooms["Room_ID"] = rooms["Room_ID"].astype(str).str.strip()
cls = rooms[rooms["Room_ID"].str.startswith('C')].copy()
labs = rooms[rooms["Room_ID"].str.startswith('L')].copy()

def s(v):
    if v is None: return ""
    if isinstance(v, float) and pd.isna(v): return ""
    return str(v).strip()

def is_combined_flag(c):
    try:
        return int(float(c.get("Is_Combined", 0))) == 1
    except Exception:
        return False

COMBINED_COURSE_CODES = {
    s(c.get("Course_Code","")).strip().upper()
    for c in (coursesCSE + coursesECE + coursesDSAI)
    if is_combined_flag(c)
}

def to_int_or_none(v):
    try:
        if v is None:
            return None
        if isinstance(v, float) and pd.isna(v):
            return None
        iv = int(float(str(v).strip()))
        if iv <= 0:
            return None
        return iv
    except Exception:
        return None

def ltp(sv):
    try:
        p = [x.strip() for x in sv.split("-")]
    except Exception:
        return [0,0,0,0,0]
    while len(p) < 5:
        p.append("0")
    return list(map(int, p[:5]))

pat = re.compile(r"^[A-Z]{1,5}\d{0,3}([+/\\-][A-Z]{1,5}\d{0,3})*$", re.I)
def valid(c):
    codes, err = [], []
    for x in c:
        code = s(x.get("Course_Code", ""))
        if not code: continue
        if code.upper() in {"NEW", "ELECTIVE"}:
            codes.append(code.upper()); continue
        if not pat.match(code):
            err.append(code)
        codes.append(code.upper())
    dup = {x for x in codes if codes.count(x) > 1 and x not in {"NEW", "ELECTIVE"}}
    if dup: err += list(dup)
    return err
def is_combined_course(code):
    try:
        return str(code).strip().upper() in COMBINED_COURSE_CODES
    except Exception:
        return False
lab_prefix_for_class_prefix = {
    "C1": "L1",
    "C2": "L2",
    "C3": "L3",
    "C4": "L4",
}

def room_meets_capacity(room_id, min_capacity):
    if min_capacity is None:
        return True
    try:
        min_capacity = float(min_capacity)
    except Exception:
        return True
    if room_id is None:
        return True
    rid = str(room_id).strip()
    if not rid:
        return True
    try:
        row = rooms[rooms["Room_ID"].astype(str).str.strip().str.upper() == rid.upper()]
        if row.empty:
            return True
        cap = pd.to_numeric(row.iloc[0].get("Capacity"), errors="coerce")
        if pd.isna(cap):
            return True
        return float(cap) >= min_capacity
    except Exception:
        return True

def room_candidates(lab=False, prefix=None, lab_prefix=None, min_capacity=None):
    df = labs if lab else cls
    if df.empty:
        return []
    cand = df.copy()
    if min_capacity is not None:
        try:
            cand = cand[pd.to_numeric(cand["Capacity"], errors="coerce").fillna(0) >= float(min_capacity)]
        except Exception:
            pass
    if prefix:
        c = cand[cand['Room_ID'].str.upper().str.startswith(prefix.upper())]
        if not c.empty:
            cand = c
        else:
            cand = df.copy()
            if min_capacity is not None:
                try:
                    cand = cand[pd.to_numeric(cand["Capacity"], errors="coerce").fillna(0) >= float(min_capacity)]
                except Exception:
                    pass
    if lab and lab_prefix:
        c = cand[cand['Room_ID'].str.upper().str.startswith(lab_prefix.upper())]
        if not c.empty:
            cand = c
    return cand["Room_ID"].tolist()

def pick_room_with_capacity_fallback(lab, day, slots_to_use, room_busy, class_prefix=None, lab_prefix=None, min_capacity=None, rr_state_key=None, rr_state=None):
    candidates = room_candidates(lab=lab, prefix=class_prefix, lab_prefix=lab_prefix, min_capacity=min_capacity)
    room = pick_room_for_slots(candidates, day, slots_to_use, room_busy, rr_state_key=rr_state_key, rr_state=rr_state)
    if room is None and min_capacity is not None:
        candidates = room_candidates(lab=lab, prefix=class_prefix, lab_prefix=lab_prefix, min_capacity=None)
        room = pick_room_for_slots(candidates, day, slots_to_use, room_busy, rr_state_key=rr_state_key, rr_state=rr_state)
    return room

def pick_room_for_slots(candidates, day, slots_to_use, room_busy, rr_state_key=None, rr_state=None):
    if not candidates:
        return None
    ordered = candidates
    if rr_state is not None and rr_state_key is not None and len(candidates) > 0:
        idx = rr_state.get(rr_state_key, 0) % len(candidates)
        ordered = candidates[idx:] + candidates[:idx]
    for cand in ordered:
        used = room_busy.get(day, {}).get(cand, set())
        if not (set(slots_to_use) & used):
            if rr_state is not None and rr_state_key is not None and len(candidates) > 0:
                rr_state[rr_state_key] = (rr_state.get(rr_state_key, 0) + 1) % len(candidates)
            return cand
    return None

def free(tt, d, ex=False):
    fb, b = [], []
    for s_ in slot_keys:
        if s_ in HARD_FORBIDDEN_SLOTS:
            if b:
                fb.append(b); b = []
            continue
        if not ex and s_ in FORBIDDEN_SLOTS:
            if b:
                fb.append(b); b = []
            continue
        if tt.at[d, s_] == "":
            b.append(s_)
        else:
            if b:
                fb.append(b); b = []
    if b: fb.append(b)
    return fb

def exact_free_blocks(tt, d, duration, ex=False):
    """
    Return contiguous blocks of free slots on day d whose total duration == duration.
    """
    blocks = []
    n = len(slot_keys)
    for i in range(n):
        total = 0.0
        cur = []
        for j in range(i, n):
            s_ = slot_keys[j]
            if s_ in HARD_FORBIDDEN_SLOTS:
                break
            if not ex and s_ in FORBIDDEN_SLOTS:
                break
            if tt.at[d, s_] != "":
                break
            cur.append(s_)
            total += slot_dur[s_]
            if abs(total - duration) < 1e-9:
                blocks.append(cur.copy())
                break
            if total > duration + 1e-9:
                break
    return blocks

def alloc_specific(tt, busy, rm, room_busy, day, slots_to_use, f, code, typ, elec, labsd, course_usage,
                   class_prefix=None, rr_state=None, hide_c004=False, skip_usage_check=False, ex=False, year_tag=None,
                   basket_used=None, basket_key=None, faculty_tt=None, semester_half=None,
                   faculty_busy_global=None, student_count=None, allow_extra_same_day=False):
    basket_num = _basket_code_parts(code) if elec else None
    for s_ in slots_to_use:
        if s_ in HARD_FORBIDDEN_SLOTS:
            return False
        if s_ not in slot_keys or tt.at[day, s_] != "":
            return False

    if code not in course_usage[day]:
        course_usage[day][code] = {"L":0,"T":0,"P":0}

    usage = course_usage[day][code]

    # For electives: Do NOT treat P as a real lab hour
    if not skip_usage_check:
        if typ == "P" and elec:
            # elective lab behaves like theory — allow unlimited placement
            pass
        else:
            if typ == "P":
                if usage["P"] >= 1:
                    return False
            else:
                if (usage["L"] + usage["T"]) >= 1 and not allow_extra_same_day:
                    return False


    r = None
    if basket_num:
        r = None
    elif not elec:
        key = (code, typ)
        if key in rm:
            candidate = rm[key]
            # if candidate is C004 we still need to check cross-branch occupancy below
            if candidate != "C004":
                used = room_busy.get(day, {}).get(candidate, set())
                if set(slots_to_use) & used:
                    return False
            if room_meets_capacity(candidate, student_count):
                r = candidate
            else:
                r = None
        if r is None:
            if typ == "P":
                lab_pref = lab_prefix_for_class_prefix.get(class_prefix, None)
                r = pick_room_with_capacity_fallback(True, day, slots_to_use, room_busy, class_prefix=None, lab_prefix=lab_pref, min_capacity=student_count, rr_state_key=class_prefix, rr_state=None)
            else:
                r = pick_room_with_capacity_fallback(False, day, slots_to_use, room_busy, class_prefix=class_prefix, lab_prefix=None, min_capacity=student_count, rr_state_key=class_prefix, rr_state=None)
            if r is None:
                return False
            rm[key] = r
        # Allow over-capacity rooms as a last-resort fallback

    else:
        # Assign rooms for electives/baskets too (avoid room clashes)
        if typ == "P":
            lab_pref = lab_prefix_for_class_prefix.get(class_prefix, None)
            r = pick_room_with_capacity_fallback(True, day, slots_to_use, room_busy, class_prefix=None, lab_prefix=lab_pref, min_capacity=student_count, rr_state_key=lab_pref, rr_state=None)
        else:
            r = pick_room_with_capacity_fallback(False, day, slots_to_use, room_busy, class_prefix=class_prefix, lab_prefix=None, min_capacity=student_count, rr_state_key=class_prefix, rr_state=None)
        if r is None:
            return False


    # Prevent cross-year basket collision (allow same-year sharing)
    if elec and basket_used is not None and basket_key and year_tag is not None:
        by_year = basket_used.setdefault(basket_key, {})
        for other_year, used in by_year.items():
            if other_year == year_tag:
                continue
            for s_ in slots_to_use:
                if (day, s_) in used:
                    return False

    # Global faculty clash check
    fac_list = split_faculty_names(f) if f else []
    if fac_list and faculty_busy_global is not None:
        for fac in fac_list:
            if set(slots_to_use) & faculty_busy_global.get(day, {}).get(fac, set()):
                return False

    # Commit the allocation to tt
    for s_ in slots_to_use:
        if is_combined_course(code):
            if hide_c004:
                if typ == "P":
                    v = f"{code} (Lab)"
                elif typ == "T":
                    v = f"{code} TUT"
                else:
                    v = f"{code}"
            else:
                if typ == "P":
                    v = f"{code} (Lab)"
                elif typ == "T":
                    v = f"{code} TUT (C004)"
                else:
                    v = f"{code} (C004)"
        else:
            if basket_num:
                rooms = get_basket_room_list(year_tag, basket_num)
                room_txt = ", ".join(rooms)
                if typ == "T":
                    v = f"{code} TUT ({room_txt})" if room_txt else f"{code} TUT"
                elif typ == "P":
                    v = f"{code} (Lab-{room_txt})" if room_txt else f"{code} (Lab)"
                else:
                    v = f"{code} ({room_txt})" if room_txt else code
            elif r:
                if elec and typ == "P":
                    v = f"{code} (Lab-{r})"
                elif typ == "T":
                    v = f"{code} TUT ({r})"
                elif typ == "P":
                    v = f"{code} (Lab-{r})"
                else:
                    v = f"{code} ({r})"
            else:
                if elec and typ == "P":
                    v = f"{code}(Lab)"
                elif typ == "T":
                    v = f"{code} TUT"
                else:
                    v = code
        tt.at[day, s_] = v
        if faculty_tt is not None and fac_list:
            if semester_half in (1, 2):
                for fac in fac_list:
                    faculty_tt.setdefault(semester_half, {}).setdefault(fac, {}).setdefault(day, {})[s_] = v

    if fac_list:
        for fac in fac_list:
            busy[day].setdefault(fac, set()).update(slots_to_use)
            if faculty_busy_global is not None:
                faculty_busy_global.setdefault(day, {}).setdefault(fac, set()).update(slots_to_use)
    if r:
        room_busy.setdefault(day, {}).setdefault(r, set()).update(slots_to_use)
    if typ == "P":
        labsd.add(day)
    course_usage[day][code][typ] += 1


    if elec and basket_used is not None and basket_key and year_tag is not None:
        by_year = basket_used.setdefault(basket_key, {})
        used = by_year.setdefault(year_tag, set())
        for s_ in slots_to_use:
            used.add((day, s_))

    return True

def alloc(tt, busy, rm, room_busy, d, f, code, h, typ="L", elec=False, labsd=set(), ex=False,
          preferred_slots=None, course_usage=None, class_prefix=None, rr_state=None, hide_c004=False,year_tag=None,
          basket_used=None, basket_key=None, faculty_tt=None, semester_half=None,
          faculty_busy_global=None, student_count=None, allow_extra_same_day=False):
    if course_usage is None:
        course_usage = {dd:{} for dd in days}
    if code not in course_usage[d]:
        course_usage[d][code] = {"L":0,"T":0,"P":0}

    usage = course_usage[d][code]

    if typ == "P":
        if usage["P"] >= 1:
            return False
    else:
        if (usage["L"] + usage["T"]) >= 1 and not allow_extra_same_day:
            return False

    if preferred_slots:
        pref_day, pref_slots = preferred_slots
        if pref_day == d:
            total = sum(slot_dur[s] for s in pref_slots)
            if abs(total - h) < 1e-9:
                if alloc_specific(tt, busy, rm, room_busy, pref_day, pref_slots, f, code, typ, elec, labsd, course_usage, class_prefix=class_prefix, rr_state=None, hide_c004=hide_c004, year_tag=year_tag, basket_used=basket_used, basket_key=basket_key, faculty_tt=faculty_tt, semester_half=semester_half, faculty_busy_global=faculty_busy_global, student_count=student_count, allow_extra_same_day=allow_extra_same_day):
                    return True

    # For L/T/P, only use exact contiguous blocks (no splitting)
    for use in exact_free_blocks(tt, d, h, ex):
        if any(s_ in HARD_FORBIDDEN_SLOTS for s_ in use): continue
        if not ex and any(s_ in FORBIDDEN_SLOTS for s_ in use): continue
        fac_list = split_faculty_names(f) if f else []
        if fac_list:
            if any(fac in busy[d] and (set(use) & busy[d][fac]) for fac in fac_list):
                continue
            if faculty_busy_global is not None:
                if any(set(use) & faculty_busy_global.get(d, {}).get(fac, set()) for fac in fac_list):
                    continue

        basket_num = _basket_code_parts(code) if elec else None
        if basket_num:
            r = None
        elif not elec:
            key = (code, typ)
            if key in rm:
                r = rm[key]
                if r != "C004":
                    used = room_busy.get(d, {}).get(r, set())
                    if set(use) & used:
                        continue
                if not room_meets_capacity(r, student_count):
                    continue
            else:
                if typ == "P":
                    lab_pref = lab_prefix_for_class_prefix.get(class_prefix, None)
                    candidates = room_candidates(lab=True, prefix=None, lab_prefix=lab_pref, min_capacity=student_count)
                    r = pick_room_for_slots(candidates, d, use, room_busy, rr_state_key=lab_pref, rr_state=None)
                else:
                    candidates = room_candidates(lab=False, prefix=class_prefix, lab_prefix=None, min_capacity=student_count)
                    r = pick_room_for_slots(candidates, d, use, room_busy, rr_state_key=class_prefix, rr_state=None)

                if r is None:
                    continue
                rm[(code, typ)] = r
        else:
            if typ == "P":
                lab_pref = lab_prefix_for_class_prefix.get(class_prefix, None)
                candidates = room_candidates(lab=True, prefix=None, lab_prefix=lab_pref, min_capacity=student_count)
                r = pick_room_for_slots(candidates, d, use, room_busy, rr_state_key=lab_pref, rr_state=None)
            else:
                candidates = room_candidates(lab=False, prefix=class_prefix, lab_prefix=None, min_capacity=student_count)
                r = pick_room_for_slots(candidates, d, use, room_busy, rr_state_key=class_prefix, rr_state=None)
            if r is None:
                continue
        if r and not room_meets_capacity(r, student_count):
            continue


        # Prevent cross-year basket collision (allow same-year sharing)
        if elec and basket_used is not None and basket_key and year_tag is not None:
            by_year = basket_used.setdefault(basket_key, {})
            conflict = False
            for other_year, used in by_year.items():
                if other_year == year_tag:
                    continue
                if any((d, s_) in used for s_ in use):
                    conflict = True
                    break
            if conflict:
                continue

        # commit allocation to cells
        for s_ in use:
            if is_combined_course(code):
                if hide_c004:
                    if typ == "P":
                        v = f"{code}(Lab)"
                    elif typ == "T":
                        v = f"{code} TUT"
                    else:
                        v = f"{code}"
                else:
                    if typ == "P":
                        v = f"{code} (Lab)"
                    elif typ == "T":
                        v = f"{code} TUT (C004)"
                    else:
                        v = f"{code} (C004)"
            else:
                if basket_num:
                    rooms = get_basket_room_list(year_tag, basket_num)
                    room_txt = ", ".join(rooms)
                    if typ == "T":
                        v = f"{code} TUT ({room_txt})" if room_txt else f"{code} TUT"
                    elif typ == "P":
                        v = f"{code} (Lab-{room_txt})" if room_txt else f"{code} (Lab)"
                    else:
                        v = f"{code} ({room_txt})" if room_txt else code
                elif r:
                    if elec and typ == "P":
                        v = f"{code} (Lab-{r})"
                    elif typ == "T":
                        v = f"{code} TUT ({r})"
                    elif typ == "P":
                        v = f"{code} (Lab-{r})"
                    else:
                        v = f"{code} ({r})"
                else:
                    if elec and typ == "P":
                        v = f"{code}(Lab)"
                    elif typ == "T":
                        v = f"{code} TUT"
                    else:
                        v = code
            tt.at[d, s_] = v
            if faculty_tt is not None and f:
                if semester_half in (1, 2):
                    faculty_tt.setdefault(semester_half, {}).setdefault(f, {}).setdefault(d, {})[s_] = v

        if f:
            busy[d].setdefault(f, set()).update(use)
        if r:
            room_busy.setdefault(d, {}).setdefault(r, set()).update(use)
        if typ == "P":
            labsd.add(d)
        course_usage[d][code][typ] += 1


        if elec and basket_used is not None and basket_key and year_tag is not None:
            by_year = basket_used.setdefault(basket_key, {})
            used = by_year.setdefault(year_tag, set())
            for s_ in use:
                used.add((d, s_))

        return True

    return False


def get_all_valid_free_slots(tt):
    valid = []
    for d in days:
        for s_ in slot_keys:
            if s_ in HARD_FORBIDDEN_SLOTS: continue
            if s_ in FORBIDDEN_SLOTS: continue
            if tt.at[d, s_] == "": valid.append((d, s_))
    return valid


def extract_contiguous_blocks(slot_list):
    # Group truly contiguous slots (by slot_keys order) per day
    by_day = {}
    slot_idx = {s:i for i,s in enumerate(slot_keys)}
    for d, s_ in slot_list:
        if s_ not in slot_idx:
            continue
        by_day.setdefault(d, []).append(slot_idx[s_])

    blocks = []
    for d in days:
        if d not in by_day:
            continue
        idxs = sorted(set(by_day[d]))
        cur = [idxs[0]]
        for i in idxs[1:]:
            if i == cur[-1] + 1:
                cur.append(i)
            else:
                blocks.append((d, [slot_keys[j] for j in cur]))
                cur = [i]
        if cur:
            blocks.append((d, [slot_keys[j] for j in cur]))
    return blocks

def try_allocate_chunk_from_block(
    tt, busy, rm, room_busy, labsd, course_usage,
    code, faculty, typ, need, day, slots,
    class_prefix=None, rr_state=None, hide_c004=False,
    allow_excluded=False, basket_used=None, basket_key=None,
    faculty_tt=None, semester_half=None, faculty_busy_global=None,
    student_count=None, exact=False
):
    n = len(slots)

    best_sub = None
    best_i = best_j = None
    best_total = None

    # 1️⃣ Find best contiguous slice
    for i in range(n):
        accum = 0.0
        sub = []          # ✅ MUST be here

        for j in range(i, n):
            s_key = slots[j]

            # 🚫 Hard-forbid excluded slots
            if s_key in HARD_FORBIDDEN_SLOTS:
                break
            if s_key in FORBIDDEN_SLOTS and not allow_excluded:
                break

            sub.append(s_key)
            accum += slot_dur[s_key]

            if accum + 1e-9 >= need:
                if exact and abs(accum - need) > 1e-9:
                    break
                if all(tt.at[day, s_] == "" for s_ in sub):
                    if best_sub is None or accum < best_total - 1e-9:
                        best_sub = list(sub)
                        best_i, best_j, best_total = i, j, accum
                break

    if best_sub is None:
        return None, None

    # 2️⃣ Allocate chosen slice
    ok = alloc_specific(
        tt, busy, rm, room_busy,
        day, best_sub,
        faculty, code, typ,
        False, labsd, course_usage,
        class_prefix=class_prefix,
        rr_state=None,
        hide_c004=hide_c004,
        basket_used=basket_used,
        basket_key=basket_key,
        faculty_tt=faculty_tt,
        semester_half=semester_half,
        faculty_busy_global=faculty_busy_global,
        student_count=student_count
    )

    if not ok:
        return None, None

    new_slots = slots[:best_i] + slots[best_j + 1:]
    return new_slots, best_sub

def assign_combined_precise_durations(
    tt, busy, rm, room_busy, labsd, course_usage, combined_core,
    rr_state=None, hide_c004=False,
    combined_sync=None, year_tag=None, semester_half=None, faculty_tt=None,
    faculty_busy_global=None
):
    ALLOWED_LECTURE_CHUNKS = [1.5, 1.0]
    if not combined_core:
        return []

    combined_list = []
    chunks_map = {}

    # ---------- BUILD CHUNKS ----------
    for c in combined_core:
        code = s(c.get("Course_Code", ""))
        if not code:
            continue

        rm[(code, "L")] = "C004"
        rm[(code, "T")] = "C004"
        rm[(code, "P")] = "C004"

        L, T, P, _, _ = ltp(c.get("L-T-P-S-C", "0-0-0-0-0"))

        ch = []

        rem = float(L)
        if abs(rem - 2.0) < 1e-9:
            # Priority order for L=2
            for a in (1.5, 0.5):
                ch.append((a, "L"))
        elif abs(rem - 1.0) < 1e-9:
            ch.append((1.0, "L"))
        else:
            while rem > 1e-9:
                if rem >= 1.5 and (abs(rem - 1.5) < 1e-9 or rem - 1.5 >= 0.5 - 1e-9):
                    ch.append((1.5, "L"))
                    rem -= 1.5
                elif rem >= 1.0 and (abs(rem - 1.0) < 1e-9 or rem - 1.0 >= 0.5 - 1e-9):
                    ch.append((1.0, "L"))
                    rem -= 1.0
                elif rem >= 0.5 - 1e-9:
                    ch.append((0.5, "L"))
                    rem -= 0.5
                else:
                    break

        rem = float(T)
        while rem > 1e-9:
            ch.append((1.0, "T"))
            rem -= 1.0

        rem = float(P)
        while rem > 1e-9:
            if rem >= 2.0:
                ch.append((2.0, "P"))
                rem -= 2.0
            elif rem >= 1.5:
                ch.append((1.5, "P"))
                rem -= 1.5
            else:
                ch.append((1.0, "P"))
                rem -= 1.0

        chunks_map[code] = sorted(ch, key=lambda x: -x[0])
        combined_list.append((code, c))

    combined_list.sort(key=lambda x: x[0])
    placed_codes = []

    # ---------- PLACEMENT ----------
    for code, c in combined_list:
        faculty = s(c.get("Faculty", ""))
        student_count = to_int_or_none(c.get("total_students"))
        chunks = chunks_map[code]

        sync_key = None
        existing_sync = None
        if combined_sync is not None and year_tag is not None:
            sync_key = (year_tag, code)
            existing_sync = combined_sync.get(sync_key)

        new_sync_entries = []
        days_used = set()
        course_ok = True

        for idx, (need, typ) in enumerate(chunks):
            allocated = False

            # ===== MIRROR LOGIC FIRST (for all types, including L) =====
            if existing_sync is not None and idx < len(existing_sync):
                sync_typ, sync_day, sync_slots = existing_sync[idx]
                if sync_typ == typ:
                    ok = alloc_specific(
                        tt, busy, rm, room_busy,
                        sync_day, sync_slots,
                        None,
                        code, typ, False, labsd, course_usage,
                        class_prefix="C0",
                        rr_state=None,
                        hide_c004=hide_c004,
                        faculty_tt=faculty_tt,
                        semester_half=semester_half,
                        faculty_busy_global=faculty_busy_global,
                        student_count=student_count
                    )
                    if ok:
                        allocated = True
                        days_used.add(sync_day)
                        continue

            # ===== NORMAL BLOCK SEARCH =====
            if not allocated:
                valid_slots = get_all_valid_free_slots(tt)
                valid_blocks = extract_contiguous_blocks(valid_slots)

                for day, slots in valid_blocks:
                    if day in days_used:
                        continue

                    _, used_slots = try_allocate_chunk_from_block(
                        tt, busy, rm, room_busy, labsd, course_usage,
                        code, faculty, typ, need, day, slots,
                        class_prefix="C0",
                        rr_state=None,
                        hide_c004=hide_c004,
                        faculty_tt=faculty_tt,
                        semester_half=semester_half,
                        faculty_busy_global=faculty_busy_global,
                        student_count=student_count,
                        exact=True
                    )

                    if used_slots is not None:
                        new_sync_entries.append((typ, day, used_slots))
                        allocated = True
                        days_used.add(day)
                        break

            if not allocated:
                course_ok = False
                break

        if course_ok and sync_key and combined_sync is not None and existing_sync is None and new_sync_entries:
            combined_sync[sync_key] = new_sync_entries

        if course_ok:
            placed_codes.append(code)

    return placed_codes

color_avail = colors.copy()
color_map = {}
legend_room_map = {}
basket_course_room_map = {}
basket_room_list_map = {}
basket_room_busy = None
ELECTIVE_SYNC_BY_YEAR = {}
GLOBAL_ROOM_BUSY = None

def reset_color_palette():
    global color_avail, color_map
    # Keep a consistent palette order across semesters
    color_avail = colors.copy()
    color_map = {}

def extract_course_code(cell_value):
    if cell_value is None:
        return ""
    val = str(cell_value).strip()
    if not val:
        return ""
    raw_course_name = val.split()[0] if val.split() else val
    raw_course_name = (
        raw_course_name.replace("T", "")
        .replace("(", "")
        .strip()
        .upper()
    )
    return raw_course_name
def get_color_for_course(course_code):
    k = course_code.strip().upper()
    if k == "": return None
    if k not in color_map:
        if color_avail: color_map[k] = color_avail.pop()
        else: color_map[k] = "CCCCCC"
    return color_map[k]

def merge_and_color(ws, courses):
    sc = 2
    mc = ws.max_column
    mr = ws.max_row

    valid_course_codes = {
        s(x.get("Course_Code", "")).replace("T", "").strip().upper()
        for x in courses
        if s(x.get("Course_Code", ""))
    }
    valid_course_codes |= {f"ELECTIVE{i}" for i in range(1, 60)}

    # Identify legend rows to avoid merging within legend tables
    legend_rows = set()
    r = 1
    while r <= mr:
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().startswith("Legend -"):
            # include title row and subsequent legend table rows
            rr = r
            while rr <= mr:
                # stop at a completely blank row
                any_val = False
                for c in range(1, mc + 1):
                    cell_v = ws.cell(rr, c).value
                    if cell_v not in (None, ""):
                        any_val = True
                        break
                if not any_val:
                    break
                legend_rows.add(rr)
                rr += 1
            r = rr
            continue
        r += 1

    # Header styling
    for col in range(2, mc + 1):
        cell = ws.cell(2, col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin

    # Merge contiguous identical entries in each row
    for r in range(3, mr + 1):
        if r in legend_rows:
            for c in range(sc, mc + 1):
                ws.cell(r, c).border = thin
            continue
        c = sc
        while c <= mc:
            raw = ws.cell(r, c).value
            if raw is None or str(raw).strip() == "":
                ws.cell(r, c).border = thin
                c += 1
                continue

            val = str(raw).strip()
            merge_cols = [c]

            # extend to all immediately-adjacent cells with same text
            next_col = c + 1
            while next_col <= mc:
                next_raw = ws.cell(r, next_col).value
                next_val = str(next_raw).strip() if next_raw is not None else ""
                if next_val == val:
                    merge_cols.append(next_col)
                    next_col += 1
                else:
                    break

            # actually merge the block if it spans >1 column
            if len(merge_cols) > 1:
                ws.merge_cells(
                    start_row=r,
                    start_column=merge_cols[0],
                    end_row=r,
                    end_column=merge_cols[-1]
                )

            # styling + colour
            cell = ws.cell(r, merge_cols[0])
            cell.alignment = Alignment(horizontal="center",
                                       vertical="center",
                                       wrap_text=True)
            cell.font = Font(bold=True)

            raw_course_name = extract_course_code(val)
            fill_color = (
                get_color_for_course(raw_course_name)
                if (raw_course_name in valid_course_codes
                    or raw_course_name.startswith("ELECTIVE"))
                else None
            )

            for cc_ in merge_cols:
                cell_ref = ws.cell(r, cc_)
                cell_ref.border = thin
                cell_ref.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                cell_ref.font = Font(bold=True)
                if fill_color:
                    cell_ref.fill = PatternFill(
                        start_color=fill_color,
                        end_color=fill_color,
                        fill_type="solid"
                    )

            c = merge_cols[-1] + 1

    # auto column widths
    for col in ws.columns:
        maxl = 0
        cl = col[0].column_letter
        for cell in col:
            v = cell.value
            if v is None:
                continue
            maxl = max(maxl, len(str(v)))
        ws.column_dimensions[cl].width = min(maxl + 2 if maxl else 8, 60)

def combined_label(c):
    if not is_combined_flag(c):
        return "No"
    dept = s(c.get("Departments","")).upper()
    sec = s(c.get("Section","")).upper()
    if dept == "CSE":
        if sec == "A":
            return "Yes, CSEB"
        if sec == "B":
            return "Yes, CSEA"
    if dept == "DSAI":
        return "Yes, ECE"
    if dept == "ECE":
        return "Yes, DSAI"
    return "Yes"

def legend_class_prefix(dept, year_tag):
    d = str(dept or "").strip().upper()
    try:
        y = int(float(year_tag))
    except Exception:
        y = None
    if d == "CSE":
        if y == 1:
            return "C1"
        if y == 3:
            return "C2"
        if y in (5, 7):
            return "C3"
    if d == "DSAI":
        if y == 1:
            return "C1"
        if y in (3, 5):
            return "C4"
    if d == "ECE":
        return "C4"
    return None

def _ensure_basket_room_busy():
    global basket_room_busy
    if basket_room_busy is not None:
        return
    if GLOBAL_ROOM_BUSY is None:
        basket_room_busy = {d: {} for d in days}
        return
    # deep copy: day -> room -> set(slots)
    basket_room_busy = {}
    for d in days:
        basket_room_busy[d] = {}
        for room, used in GLOBAL_ROOM_BUSY.get(d, {}).items():
            basket_room_busy[d][room] = set(used)

def _basket_code_parts(code):
    try:
        parts = str(code).strip().split()
        if len(parts) >= 3 and parts[0].upper() == "ELECTIVE" and parts[1].upper() == "BASKET":
            return parts[2]
        if len(parts) >= 1 and parts[0].upper() == "ELECTIVE":
            return "ELECTIVE"
    except Exception:
        pass
    return None

def get_basket_room_list(year_tag, basket):
    if year_tag is None:
        return []
    return basket_room_list_map.get((int(year_tag), str(basket)), [])

def assign_basket_rooms_for_group(year_tag, basket, course_group):
    if year_tag is None:
        return
    basket = str(basket)
    basket = normalize_elective_basket(year_tag, basket)
    if not basket or basket == "0":
        return
    sync = ELECTIVE_SYNC_BY_YEAR.get(int(year_tag), {})
    sync_key = f"Y{int(year_tag)}_B{basket}"
    pref = sync.get(sync_key)
    if not pref:
        return
    if isinstance(pref, list):
        blocks = [(d, s) for d, s in pref]
    else:
        day = pref.get("day")
        slots = pref.get("slots", [])
        blocks = [(day, slots)]
    # filter invalid
    blocks = [(d, s) for d, s in blocks if d and s]
    if not blocks:
        return

    _ensure_basket_room_busy()
    # Prepare course list with metadata
    course_items = []
    for c in course_group:
        code = s(c.get("Course_Code","")).strip().upper()
        if not code:
            continue
        student_count = to_int_or_none(c.get("total_students"))
        class_prefix = legend_class_prefix(c.get("Departments",""), year_tag)
        course_items.append((code, student_count, class_prefix))

    rooms_list = []
    # Assign rooms per block independently (no cross-block coupling)
    for bd, bslots in blocks:
        # track rooms used in this block
        used_rooms = set(basket_room_busy.get(bd, {}).keys())
        # assign harder courses first (higher student count)
        sorted_items = sorted(course_items, key=lambda x: (-(x[1] or 0), x[0]))
        for code, student_count, class_prefix in sorted_items:
            map_key = (int(year_tag), basket, code)
            # build candidates for this course
            candidates = room_candidates(lab=False, prefix=class_prefix, lab_prefix=None, min_capacity=student_count)
            room = None
            for cand in candidates:
                if cand in used_rooms:
                    continue
                used = basket_room_busy.get(bd, {}).get(cand, set())
                if set(bslots) & used:
                    continue
                room = cand
                break
            if room is None and student_count is not None:
                candidates = room_candidates(lab=False, prefix=class_prefix, lab_prefix=None, min_capacity=None)
                for cand in candidates:
                    if cand in used_rooms:
                        continue
                    used = basket_room_busy.get(bd, {}).get(cand, set())
                    if set(bslots) & used:
                        continue
                    room = cand
                    break
            if room is None:
                # no room available for this course in this block
                continue
            used_rooms.add(room)
            basket_room_busy.setdefault(bd, {}).setdefault(room, set()).update(bslots)
            rooms_list.append(room)
            # store per-course rooms
            rooms_for_course = basket_course_room_map.get(map_key, [])
            if room not in rooms_for_course:
                rooms_for_course.append(room)
            basket_course_room_map[map_key] = rooms_for_course
    # normalize basket room list
    uniq = []
    for r in rooms_list:
        if r not in uniq:
            uniq.append(r)
    basket_room_list_map[(int(year_tag), basket)] = uniq

def assign_basket_course_room(c):
    if not is_combined_flag(c):
        pass
    if s(c.get("Elective","")) != "1":
        return ""
    basket = s(c.get("ElectiveBasket","0"))
    try:
        year_tag = int(float(s(c.get("Semester",""))))
    except Exception:
        year_tag = None
    if year_tag is None:
        return ""
    basket = normalize_elective_basket(year_tag, basket)
    if not basket or basket == "0":
        return ""
    sync = ELECTIVE_SYNC_BY_YEAR.get(year_tag, {})
    sync_key = f"Y{year_tag}_B{basket}"
    pref = sync.get(sync_key)
    if not pref:
        return ""
    if isinstance(pref, list):
        blocks = [(d, s) for d, s in pref]
    else:
        day = pref.get("day")
        slots = pref.get("slots", [])
        blocks = [(day, slots)]
    blocks = [(d, s) for d, s in blocks if d and s]
    if not blocks:
        return ""
    code = s(c.get("Course_Code","")).strip().upper()
    if not code:
        return ""
    map_key = (year_tag, basket, code)
    if map_key in basket_course_room_map:
        rooms = basket_course_room_map[map_key]
        return ", ".join(rooms) if rooms else ""

    _ensure_basket_room_busy()
    student_count = to_int_or_none(c.get("total_students"))
    class_prefix = legend_class_prefix(c.get("Departments",""), year_tag)
    candidates = room_candidates(lab=False, prefix=class_prefix, lab_prefix=None, min_capacity=student_count)
    rooms_for_course = []
    for bd, bslots in blocks:
        room = None
        for cand in candidates:
            used = basket_room_busy.get(bd, {}).get(cand, set())
            if set(bslots) & used:
                continue
            room = cand
            break
        if room is None and student_count is not None:
            candidates = room_candidates(lab=False, prefix=class_prefix, lab_prefix=None, min_capacity=None)
            for cand in candidates:
                used = basket_room_busy.get(bd, {}).get(cand, set())
                if set(bslots) & used:
                    continue
                room = cand
                break
        if room is None:
            rooms_for_course = []
            break
        rooms_for_course.append(room)
        basket_room_busy.setdefault(bd, {}).setdefault(room, set()).update(bslots)
    if not rooms_for_course:
        return ""
    uniq_rooms = []
    for r in rooms_for_course:
        if r not in uniq_rooms:
            uniq_rooms.append(r)
    basket_course_room_map[map_key] = uniq_rooms
    return ", ".join(uniq_rooms)

def add_csv_legend_block(ws, course_list, legend_title, half=None, room_map_key=None):
    ws.append([])
    ws.append([])
    ws.append([f"Legend - {legend_title}"])

    title_cell = ws.cell(row=ws.max_row, column=1)
    title_cell.font = Font(bold=True, size=13)

    headers = ["Course Code","Course Title","L-T-P-S-C","Faculty","Elective?","Combined?","Room(s)"]
    ws.append(headers)

    for i in range(1,8):
        c = ws.cell(ws.max_row,i)
        c.font = Font(bold=True)
        c.border = thin

    seen = set()
    rmap = legend_room_map.get(room_map_key, {}) if room_map_key is not None else {}
    scheduled_codes = set(rmap.keys())

    for c in course_list:
        if half in (1, 2):
            h = s(c.get("Semester_Half", "0"))
            if h not in (str(half), "0"):
                continue
        code = str(c.get("Course_Code","")).strip()
        if not code or code in seen:
            continue
        seen.add(code)

        is_elec = s(c.get("Elective","")) == "1"
        basket = s(c.get("ElectiveBasket","0"))
        if is_elec and basket and basket != "0":
            elec_label = f"Yes, {basket}"
        elif is_elec:
            elec_label = "Yes"
        else:
            elec_label = "No"
        room_str = ""
        # Prefer basket room list for electives in baskets (even if not in rmap)
        if is_elec and basket and basket != "0":
            assigned = assign_basket_course_room(c)
            if assigned:
                room_str = assigned
            else:
                try:
                    ytag = to_int_or_none(c.get("Semester",""))
                except Exception:
                    ytag = None
                if ytag is not None:
                    rooms = get_basket_room_list(ytag, normalize_elective_basket(ytag, basket))
                    if rooms:
                        room_str = ", ".join(rooms)

        is_scheduled = (not scheduled_codes) or (code in scheduled_codes) or is_combined_flag(c)
        if is_scheduled:
            if not room_str and room_map_key is not None:
                rooms = set(rmap.get(code, set()))
                room_str = ", ".join(sorted(rooms))
            # If combined course hides room in timetable, show C004 in legend
            if not room_str and is_combined_flag(c):
                room_str = "C004"

        ws.append([
            code,
            c.get("Course_Title",""),
            c.get("L-T-P-S-C",""),
            shorten_faculty_name(c.get("Faculty","")),
            elec_label,
            combined_label(c),
            room_str
        ])

    ws.append([])

def _safe_sheet_name(name, used):
    raw = str(name)
    # Excel sheet name invalid chars: : \ / ? * [ ]
    invalid = set(':/\\?*[]')
    safe = "".join(" " if ch in invalid else ch for ch in raw)
    safe = safe.replace("\n", " ").replace("\r", " ").strip()
    if not safe:
        safe = "Sheet"
    base = safe[:31]
    if base not in used:
        used.add(base)
        return base
    i = 2
    while True:
        suffix = f"_{i}"
        cand = (base[:31 - len(suffix)] + suffix)[:31]
        if cand not in used:
            used.add(cand)
            return cand
        i += 1

def write_faculty_workbook(faculty_map, filename, course_index=None):
    wb = Workbook()
    used = set()
    first = True
    reset_color_palette()
    if course_index is None:
        course_index = {}
    for faculty in sorted(faculty_map.keys()):
        short_faculty = shorten_faculty_name(faculty)
        sheet_name = _safe_sheet_name(short_faculty, used)
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)

        ws.append(["Day"] + slot_keys)
        for d in days:
            row = [d]
            for s_ in slot_keys:
                cell_val = faculty_map.get(faculty, {}).get(d, {}).get(s_, "")
                row.append(cell_val)
            ws.append(row)

        # Color-code courses per cell (same course => same color)
        for r in range(2, ws.max_row + 1):
            for c in range(2, ws.max_column + 1):
                cell = ws.cell(r, c)
                code = extract_course_code(cell.value)
                if not code:
                    continue
                fill_color = get_color_for_course(code)
                if fill_color:
                    cell.fill = PatternFill(
                        start_color=fill_color,
                        end_color=fill_color,
                        fill_type="solid"
                    )

        # Legend section
        ws.append([])
        ws.append(["Legend"])
        legend_title = ws.cell(row=ws.max_row, column=1)
        legend_title.font = Font(bold=True, size=13)
        ws.append(["Course Code", "Course Title", "Branch"])
        for i in range(1, 4):
            c = ws.cell(ws.max_row, i)
            c.font = Font(bold=True)
            c.border = thin

        # Collect only courses actually scheduled for this faculty
        used_codes = set()
        for d in days:
            for s_ in slot_keys:
                val = faculty_map.get(faculty, {}).get(d, {}).get(s_, "")
                code = extract_course_code(val)
                if code and code.upper() not in {"MINOR", "MINOR SLOTS"}:
                    used_codes.add(code)

        rows = []
        for code in sorted(used_codes):
            entries = course_index.get(code, [])
            # filter to rows matching this faculty if possible
            filtered = [e for e in entries if s(e[2]) == s(faculty)]
            pick = filtered if filtered else entries
            for title, dept, _fac in pick:
                rows.append((code, title, dept))

        for code, title, dept in sorted(set(rows), key=lambda x: (x[0], x[1], x[2])):
            ws.append([code, title, dept])
    if first:
        wb.active.title = "No Faculty"
    wb.save(filename)

def generate(courses, ws, label, seed, elective_sync,
             room_prefix=None, elective_room_map=None,
             room_busy_global=None, hide_c004=False,
             year_tag=None, combined_sync=None,semester_half=None,
             basket_used_global=None, faculty_tt=None,
             full_sem_sync=None,
             faculty_busy_global=None,
             display_slot_keys=None):
    if elective_room_map is None:
        elective_room_map = {}
    if valid(courses): return []
    
    ws.append([""]); ws.append([label])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    
    tt = pd.DataFrame("", index=days, columns=slot_keys)
    busy = {d:{} for d in days}
    
    if room_busy_global is not None:
        room_busy = room_busy_global
    else:
        room_busy = {d:{} for d in days}

    rm = {}
    labsd = set()
    course_usage = {d:{} for d in days}
    rr_state = {}

    elec = [x for x in courses if s(x.get("Elective","")) == "1"]
    combined_core = [x for x in courses if s(x.get("Elective","")) != "1" and is_combined_flag(x)]
    regular_core = [x for x in courses if s(x.get("Elective","")) != "1" and (not is_combined_flag(x))]

    baskets = {}; elec_no_baskets = []
    for e in elec:
        b = s(e.get("ElectiveBasket","0"))
        b_norm = normalize_elective_basket(year_tag, b)
        if b_norm and b_norm != "0":
            baskets.setdefault(b_norm, []).append(e)
        else:
            elec_no_baskets.append(e)
    basket_reps = []
    for b, group in sorted(baskets.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 0):
        chosen = group[0]
        # Use the max hours across all courses in this basket
        max_l = max((int(float(s(g.get("L","0")) or 0)) for g in group), default=0)
        max_t = max((int(float(s(g.get("T","0")) or 0)) for g in group), default=0)
        max_p = max((int(float(s(g.get("P","0")) or 0)) for g in group), default=0)
        max_s = max((int(float(s(g.get("S","0")) or 0)) for g in group), default=0)
        max_c = max((int(float(s(g.get("C","0")) or 0)) for g in group), default=0)
        sync_identifier = f"Y{year_tag}_B{b}" if year_tag is not None else f"B{b}"
        if year_tag == 3 and b == "ELECTIVE":
            display_code = "Elective"
            display_title = "Elective"
        else:
            display_code = f"Elective Basket {b}"
            display_title = chosen.get("Course_Title","") or chosen.get("Course_Code","")
        basket_reps.append({
            "Course_Code": display_code,
            "Course_Title": display_title,
            # keep blank to avoid cross-branch faculty clash on basket slots
            "Faculty": "",
            "L-T-P-S-C": f"{max_l}-{max_t}-{max_p}-{max_s}-{max_c}",
            "Elective": "1",
            "ElectiveBasket": b,
            "_sync_name": sync_identifier
        })

    for e in elec_no_baskets:
        basket = s(e.get("ElectiveBasket","0"))
        if basket and basket != "0":
            sync_n = f"Y{year_tag}_B{basket}" if year_tag is not None else f"B{basket}"
        else:
            sync_n = s(e.get("Course_Code"))
        e["_sync_name"] = sync_n if sync_n else None
    # Place basket slots first so all sections share the same basket time
    elec_final = basket_reps + elec_no_baskets

    for c in combined_core:
        code = s(c.get("Course_Code",""))
        rm[(code,"L")] = "C004"; rm[(code,"T")] = "C004"; rm[(code,"P")] = "C004"

    # Track hours successfully pre-placed for full-semester courses
    preplaced_hours = {}

    # Pre-place full-semester courses for second half using first-half slots
    if semester_half == 2 and full_sem_sync is not None and year_tag is not None:
        for c in courses:
            if s(c.get("Semester_Half","")) != "0":
                continue
            code = s(c.get("Course_Code",""))
            f = s(c.get("Faculty",""))
            student_count = to_int_or_none(c.get("total_students"))
            if not code:
                continue
            is_elec_flag = (code.startswith("Elective") or s(c.get("Elective","")) == "1")
            basket = s(c.get("ElectiveBasket","0"))
            basket_key = f"B{basket}" if (is_elec_flag and basket and basket != "0") else None
            fs_key = full_sem_key(c, year_tag)
            sync_for_course = full_sem_sync.get(fs_key, {})
            for typ in ("L", "T", "P"):
                for sync_day, sync_slots in sync_for_course.get(typ, []):
                    ok = alloc_specific(
                        tt, busy, rm, room_busy,
                        sync_day, sync_slots,
                        f,
                        code, typ, is_elec_flag, labsd, course_usage,
                        class_prefix=room_prefix,
                        rr_state=None,
                        hide_c004=hide_c004,
                        year_tag=year_tag,
                        basket_used=basket_used_global,
                        basket_key=basket_key,
                        faculty_tt=faculty_tt,
                        semester_half=semester_half,
                        faculty_busy_global=faculty_busy_global,
                        skip_usage_check=True,
                        student_count=student_count
                    )
                    if ok:
                        preplaced_hours.setdefault(fs_key, {}).setdefault(typ, 0.0)
                        preplaced_hours[fs_key][typ] += sum(slot_dur[s] for s in sync_slots)

    def place_course_list(course_list, start_idx_ref):
        placed_list = []
        for c in course_list:
            f = s(c.get("Faculty",""))
            code = s(c.get("Course_Code","UNKNOWN"))
            is_elec_flag = (code.startswith("Elective") or s(c.get("Elective","")) == "1")
            basket = s(c.get("ElectiveBasket","0"))
            basket_key = f"B{basket}" if (is_elec_flag and basket and basket != "0") else None
            student_count = to_int_or_none(c.get("total_students"))
            L, T, P, S, Cc = ltp(c.get("L-T-P-S-C","0-0-0-0-0"))
            is_full_sem = s(c.get("Semester_Half","")) == "0"
            fs_key = full_sem_key(c, year_tag)
            typ_counts = {"L":0,"T":0,"P":0}
            # Track lecture split mode for exact 2.0 hours (priority order handled by attempts)
            if "_L2_mode" not in c:
                c["_L2_mode"] = "1.5+0.5"

            for h, typ in [(L,"L"), (T,"T"), (P,"P")]:
                # If full-sem course was pre-placed, reduce remaining hours
                if is_full_sem and semester_half == 2 and full_sem_sync is not None:
                    pre_h = preplaced_hours.get(fs_key, {}).get(typ, 0.0)
                    h = max(0.0, h - pre_h)
                    if h <= 1e-9:
                        continue
                attempts = 0
                no_progress = 0
                while h > 1e-9 and attempts < 60 and no_progress < 8:
                    # Enforce strict durations per type:
                    # - Lecture (L) => L=2 uses priority order: 1.5+0.5, 1+1, 1+0.5+0.5, 0.5x4, 2.0
                    # - Tutorial (T) => always 1.0 hour
                    # - Practical/Lab (P) => prefer 2.0, else 1.5, else 1.0 (only if remaining h is smaller)
                    if typ == "L":
                        if abs(h - 2.0) < 1e-9:
                            # Advance to next split mode only if earlier modes failed repeatedly
                            if attempts == 10:
                                c["_L2_mode"] = "1+1"
                            elif attempts == 20:
                                c["_L2_mode"] = "1+0.5+0.5"
                            elif attempts == 30:
                                c["_L2_mode"] = "0.5x4"
                            elif attempts == 38:
                                c["_L2_mode"] = "2.0"

                            mode = c.get("_L2_mode", "1.5+0.5")
                            if mode == "1.5+0.5":
                                a = 1.5 if h >= 1.5 - 1e-9 else 0.5
                            elif mode == "1+1":
                                a = 1.0
                            elif mode == "1+0.5+0.5":
                                a = 1.0 if h >= 1.0 - 1e-9 else 0.5
                            elif mode == "0.5x4":
                                a = 0.5
                            else:
                                a = 2.0
                        elif abs(h - 1.0) < 1e-9:
                            a = 1.0
                        elif h >= 1.5 - 1e-9:
                            a = 1.5
                        elif h >= 1.0 - 1e-9:
                            a = 1.0
                        else:
                            a = 0.5
                    elif typ == "T":
                        a = 1.0
                    elif typ == "P":
                        # For labs prefer 2.0 blocks; if remaining hours < 2, allow smaller lab chunk
                        if h >= 2.0 - 1e-9:
                            a = 2.0
                        elif h >= 1.5 - 1e-9:
                            a = 1.5
                        else:
                            a = 1.0
                    else:
                        a = 1.0
                    placed = False
                    sync_name = c.get("_sync_name", None)

                    if is_elec_flag and sync_name and sync_name in elective_room_map:
                        for ttkey in [("L"), ("T"), ("P")]:
                            rm[(code, ttkey)] = elective_room_map[sync_name]

                    if sync_name and sync_name in elective_sync:
                        pref = elective_sync[sync_name]
                        if isinstance(pref, list):
                            # mirror all basket blocks
                            any_ok = False
                            for pd, pslots in pref:
                                ok = alloc_specific(
                                    tt, busy, rm, room_busy,
                                    pd, pslots,
                                    f, code, typ, is_elec_flag, labsd, course_usage,
                                    class_prefix=room_prefix,
                                    rr_state=None,
                                    hide_c004=hide_c004,
                                    year_tag=year_tag,
                                    basket_used=basket_used_global,
                                    basket_key=basket_key,
                                    faculty_tt=faculty_tt,
                                    semester_half=semester_half,
                                    faculty_busy_global=faculty_busy_global,
                                    student_count=student_count,
                                    skip_usage_check=True
                                )
                                if ok:
                                    any_ok = True
                                    h -= sum(slot_dur[s] for s in pslots)
                            placed = any_ok
                        else:
                            if alloc(tt, busy, rm, room_busy, pref["day"], f, code, a, typ, is_elec_flag, labsd, False, preferred_slots=(pref["day"], pref["slots"]), course_usage=course_usage, class_prefix=room_prefix, rr_state=None,hide_c004=hide_c004,year_tag=year_tag, basket_used=basket_used_global, basket_key=basket_key, faculty_tt=faculty_tt, semester_half=semester_half, faculty_busy_global=faculty_busy_global, student_count=student_count, allow_extra_same_day=(typ == "L" and a <= 1.0 + 1e-9)):
                                h -= a; placed = True

                    if not placed:
                        for i in range(2):
                            if is_elec_flag:
                                d_order = days[:]
                            else:
                                start_idx = start_idx_ref[0]
                                d_order = days[start_idx:] + days[:start_idx]
                                start_idx_ref[0] = (start_idx_ref[0] + 1) % len(days)
                            for d in d_order:
                                if alloc(tt, busy, rm, room_busy, d, f, code, a, typ, is_elec_flag, labsd, False, course_usage=course_usage, class_prefix=room_prefix, rr_state=None,hide_c004=hide_c004,year_tag=year_tag, basket_used=basket_used_global, basket_key=basket_key, faculty_tt=faculty_tt, semester_half=semester_half, faculty_busy_global=faculty_busy_global, student_count=student_count, allow_extra_same_day=(typ == "L" and a <= 1.0 + 1e-9)):
                                    h -= a; placed = True
                                    typ_counts[typ] = typ_counts.get(typ, 0) + 1
                                    break
                            if placed:
                                break
                    if not placed:
                        for d in days:
                            if alloc(tt, busy, rm, room_busy, d, f, code, a, typ, is_elec_flag, labsd, True, course_usage=course_usage, class_prefix=room_prefix, rr_state=None,hide_c004=hide_c004,year_tag=year_tag, basket_used=basket_used_global, basket_key=basket_key, faculty_tt=faculty_tt, semester_half=semester_half, faculty_busy_global=faculty_busy_global, student_count=student_count, allow_extra_same_day=(typ == "L" and a <= 1.0 + 1e-9)):
                                h -= a; placed = True
                                typ_counts[typ] = typ_counts.get(typ, 0) + 1
                                break

                    if placed:
                        no_progress = 0
                    else:
                        no_progress += 1
                    if placed and sync_name:
                        code_u = str(code).strip().upper()
                        if code_u.startswith("ELECTIVE BASKET") or code_u == "ELECTIVE":
                            elective_sync[sync_name] = collect_code_slot_blocks(tt, code)
                        elif sync_name not in elective_sync:
                            for dcheck in days:
                                slots_used = [s_ for s_ in slot_keys if tt.at[dcheck, s_].startswith(code)]
                                if slots_used:
                                    accum = []; acc_dur = 0.0
                                    for s_ in slots_used:
                                        accum.append(s_); acc_dur += slot_dur[s_]
                                        if acc_dur + 1e-9 >= a:
                                            elective_sync[sync_name] = {"day": dcheck, "slots": accum.copy()}
                                            break
                                    if sync_name in elective_sync: break

                    attempts += 1
            placed_list.append(c)
        return placed_list

    start_idx_ref = [seed % len(days)]
    elec_final.sort(key=lambda x: 0 if x.get("_sync_name") in elective_sync else 1)
    
    priority_placed = place_course_list(elec_final, start_idx_ref)
    # Assign rooms for each elective course in baskets (for legend + basket display)
    if year_tag is not None:
        # Ensure basket sync exists even if earlier capture missed it
        for b in baskets.keys():
            sync_id = f"Y{year_tag}_B{b}"
            if sync_id not in elective_sync:
                if year_tag == 3 and b == "ELECTIVE":
                    placeholder_code = "Elective"
                else:
                    placeholder_code = f"Elective Basket {b}"
                blocks = collect_code_slot_blocks(tt, placeholder_code)
                if blocks:
                    elective_sync[sync_id] = blocks
        for b, group in baskets.items():
            assign_basket_rooms_for_group(year_tag, b, group)
        apply_basket_rooms_to_tt(tt, year_tag)

    combined_placed = assign_combined_precise_durations(
        tt, busy, rm, room_busy, labsd, course_usage, combined_core,
        rr_state=None, hide_c004=hide_c004,  combined_sync=combined_sync, year_tag=year_tag,semester_half=semester_half, faculty_tt=faculty_tt,
        faculty_busy_global=faculty_busy_global
    )
    regular_placed = place_course_list(regular_core, start_idx_ref)

    # Label minor slots for semesters 3 and 5
    if year_tag in (3, 5):
        for d in days:
            for s_ in MINOR_SLOTS:
                if s_ in tt.columns and tt.at[d, s_] == "":
                    tt.at[d, s_] = "Minor slot"
    # Label break slots across all semesters
    for d in days:
        for s_ in BREAK_SLOTS:
            if s_ in tt.columns and tt.at[d, s_] == "":
                tt.at[d, s_] = "Break"

    if display_slot_keys is None:
        display_slot_keys = slot_keys
    # For first-half runs, record full-sem course slots after allocation
    if full_sem_sync is not None and semester_half == 1:
        build_full_sem_sync_from_tt(tt, courses, year_tag, full_sem_sync)

    # capture room mapping for legend (per block label)
    legend_room_map[label] = build_room_map_from_tt(tt)

    ws.append(["Day"] + display_slot_keys)
    for d in days:
        ws.append([d] + [tt.at[d, s] for s in display_slot_keys])
    ws.append([""])
    return (priority_placed + regular_placed + combined_core)
def split(c):
    f = [x for x in c if s(x.get("Semester_Half","")) in ["1","0"]]
    s2 = [x for x in c if s(x.get("Semester_Half","")) in ["2","0"]]
    return f, s2

if __name__ == "__main__":
    wb = Workbook()
    seed = random.randint(0, 999999)

    elective_room_map = {}
    global_room_busy = {d: {} for d in days}
    basket_used_global = {}
    faculty_tt = {1: {}, 2: {}}
    faculty_busy_global = {
        1: {d: {} for d in days},
        2: {d: {} for d in days}
    }

    sync_sem1 = {}
    sync_sem3 = {}
    sync_sem5_cse = {}
    sync_sem5_de = {}
    sync_sem7 = {}
    combined_sync_cse_sem1 = {}
    combined_sync_cse_sem3 = {}
    combined_sync_cse_sem5 = {}
    combined_sync_de_sem1 = {}
    combined_sync_de_sem3 = {}
    combined_sync_de_sem5 = {}
    full_sem_sync_sem1 = {}
    full_sem_sync_sem3 = {}
    full_sem_sync_sem5 = {}
    full_sem_sync_sem7 = {}

    # Expose for legend elective-room assignment
    GLOBAL_ROOM_BUSY = global_room_busy
    ELECTIVE_SYNC_BY_YEAR = {1: sync_sem1, 3: sync_sem3, 5: sync_sem5_cse, 7: sync_sem7}

    sem1_display_slots = [s for s in slot_keys if s not in ABSOLUTELY_FORBIDDEN_SLOTS]

    unscheduled = []

    ws1 = wb.active
    ws1.title = "CSE-I Timetable"
    cAf, cAs = split(coursesAI)
    cBf, cBs = split(coursesBI)
    
    csea_block = generate(cAf, ws1, "CSEA I First Half", seed+0, sync_sem1, room_prefix='C1', elective_room_map=elective_room_map, room_busy_global=global_room_busy,hide_c004=True,year_tag=1,combined_sync=combined_sync_cse_sem1,semester_half=1, basket_used_global=basket_used_global, faculty_tt=faculty_tt, full_sem_sync=full_sem_sync_sem1, faculty_busy_global=faculty_busy_global[1], display_slot_keys=sem1_display_slots)
    csea_block2 = generate(cAs, ws1, "CSEA I Second Half", seed+1, sync_sem1, room_prefix='C1', elective_room_map=elective_room_map, room_busy_global=global_room_busy,hide_c004=True,year_tag=1,combined_sync=combined_sync_cse_sem1,semester_half=2, basket_used_global=basket_used_global, faculty_tt=faculty_tt, full_sem_sync=full_sem_sync_sem1, faculty_busy_global=faculty_busy_global[2], display_slot_keys=sem1_display_slots)
    unscheduled += collect_unscheduled(cAf, csea_block, "CSEA I First Half", year_tag=1, elective_sync=sync_sem1)
    unscheduled += collect_unscheduled(cAs, csea_block2, "CSEA I Second Half", year_tag=1, elective_sync=sync_sem1)
    add_csv_legend_block(ws1, coursesCSEA_I, "CSEA I - First Half", half=1, room_map_key="CSEA I First Half")
    add_csv_legend_block(ws1, coursesCSEA_I, "CSEA I - Second Half", half=2, room_map_key="CSEA I Second Half")
    
    cseb_block = generate(cBf, ws1, "CSEB I First Half", seed+2, sync_sem1, room_prefix='C1', elective_room_map=elective_room_map, room_busy_global=global_room_busy,hide_c004=True,year_tag=1,combined_sync=combined_sync_cse_sem1,semester_half=1, basket_used_global=basket_used_global, faculty_tt=faculty_tt, full_sem_sync=full_sem_sync_sem1, faculty_busy_global=faculty_busy_global[1], display_slot_keys=sem1_display_slots)
    cseb_block2 = generate(cBs, ws1, "CSEB I Second Half", seed+3, sync_sem1, room_prefix='C1', elective_room_map=elective_room_map, room_busy_global=global_room_busy,hide_c004=True,year_tag=1,combined_sync=combined_sync_cse_sem1,semester_half=2, basket_used_global=basket_used_global, faculty_tt=faculty_tt, full_sem_sync=full_sem_sync_sem1, faculty_busy_global=faculty_busy_global[2], display_slot_keys=sem1_display_slots)
    unscheduled += collect_unscheduled(cBf, cseb_block, "CSEB I First Half", year_tag=1, elective_sync=sync_sem1)
    unscheduled += collect_unscheduled(cBs, cseb_block2, "CSEB I Second Half", year_tag=1, elective_sync=sync_sem1)
    add_csv_legend_block(ws1, coursesCSEB_I, "CSEB I - First Half", half=1, room_map_key="CSEB I First Half")
    add_csv_legend_block(ws1, coursesCSEB_I, "CSEB I - Second Half", half=2, room_map_key="CSEB I Second Half")

    
    combined_i_courses = (csea_block or []) + (csea_block2 or []) + (cseb_block or []) + (cseb_block2 or [])
    reset_color_palette()
    merge_and_color(ws1, combined_i_courses)

    # --- DSAI-I ---
    ws7 = wb.create_sheet("DSAI-I Timetable")
    d1f_i, d1s_i = split(coursesDSAI_I)
    dsai1_block1 = generate(d1f_i, ws7, "DSAI-I First Half", seed+16, sync_sem1, room_prefix='C1', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=1,combined_sync=combined_sync_de_sem1,semester_half=1, basket_used_global=basket_used_global, faculty_tt=faculty_tt, full_sem_sync=full_sem_sync_sem1, faculty_busy_global=faculty_busy_global[1], display_slot_keys=sem1_display_slots)
    dsai1_block2 = generate(d1s_i, ws7, "DSAI-I Second Half", seed+17, sync_sem1, room_prefix='C1', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=1,combined_sync=combined_sync_de_sem1,semester_half=2, basket_used_global=basket_used_global, faculty_tt=faculty_tt, full_sem_sync=full_sem_sync_sem1, faculty_busy_global=faculty_busy_global[2], display_slot_keys=sem1_display_slots)
    unscheduled += collect_unscheduled(d1f_i, dsai1_block1, "DSAI-I First Half", year_tag=1, elective_sync=sync_sem1)
    unscheduled += collect_unscheduled(d1s_i, dsai1_block2, "DSAI-I Second Half", year_tag=1, elective_sync=sync_sem1)
    add_csv_legend_block(ws7, coursesDSAI_I, "DSAI I - First Half", half=1, room_map_key="DSAI-I First Half")
    add_csv_legend_block(ws7, coursesDSAI_I, "DSAI I - Second Half", half=2, room_map_key="DSAI-I Second Half")

    combined_dsai1_courses = (dsai1_block1 or []) + (dsai1_block2 or [])
    merge_and_color(ws7, combined_dsai1_courses)

    # --- ECE-I ---
    ws9 = wb.create_sheet("ECE-I Timetable")
    e1f_i, e1s_i = split(coursesECE_I)
    ece1_block1 = generate(e1f_i, ws9, "ECE-I First Half", seed+20, sync_sem1, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=1,combined_sync=combined_sync_de_sem1,semester_half=1, basket_used_global=basket_used_global, faculty_tt=faculty_tt, full_sem_sync=full_sem_sync_sem1, faculty_busy_global=faculty_busy_global[1], display_slot_keys=sem1_display_slots)
    ece1_block2 = generate(e1s_i, ws9, "ECE-I Second Half", seed+21, sync_sem1, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=1,combined_sync=combined_sync_de_sem1,semester_half=2, basket_used_global=basket_used_global, faculty_tt=faculty_tt, full_sem_sync=full_sem_sync_sem1, faculty_busy_global=faculty_busy_global[2], display_slot_keys=sem1_display_slots)
    unscheduled += collect_unscheduled(e1f_i, ece1_block1, "ECE-I First Half", year_tag=1, elective_sync=sync_sem1)
    unscheduled += collect_unscheduled(e1s_i, ece1_block2, "ECE-I Second Half", year_tag=1, elective_sync=sync_sem1)
    add_csv_legend_block(ws9, coursesECE_I, "ECE I - First Half", half=1, room_map_key="ECE-I First Half")
    add_csv_legend_block(ws9, coursesECE_I, "ECE I - Second Half", half=2, room_map_key="ECE-I Second Half")

    combined_ece1_courses = (ece1_block1 or []) + (ece1_block2 or [])
    merge_and_color(ws9, combined_ece1_courses)
    # --- CSE-III (Sections A & B) ---
    ws2 = wb.create_sheet("CSE-III Timetable")
    c1f, c1s = split(coursesA); c2f, c2s = split(coursesB)
    
    csea3_block1 = generate(c1f, ws2, "CSEA III First Half", seed+4, sync_sem3, room_prefix='C2', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=3,combined_sync=combined_sync_cse_sem3,semester_half=1, basket_used_global=basket_used_global, faculty_tt=faculty_tt, full_sem_sync=full_sem_sync_sem3, faculty_busy_global=faculty_busy_global[1])
    csea3_block2 = generate(c1s, ws2, "CSEA III Second Half", seed+5, sync_sem3, room_prefix='C2', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=3,combined_sync=combined_sync_cse_sem3,semester_half=2, basket_used_global=basket_used_global, faculty_tt=faculty_tt, full_sem_sync=full_sem_sync_sem3, faculty_busy_global=faculty_busy_global[2])
    unscheduled += collect_unscheduled(c1f, csea3_block1, "CSEA III First Half", year_tag=3, elective_sync=sync_sem3)
    unscheduled += collect_unscheduled(c1s, csea3_block2, "CSEA III Second Half", year_tag=3, elective_sync=sync_sem3)
    add_csv_legend_block(ws2, coursesCSEA_III, "CSEA III - First Half", half=1, room_map_key="CSEA III First Half")
    add_csv_legend_block(ws2, coursesCSEA_III, "CSEA III - Second Half", half=2, room_map_key="CSEA III Second Half")

    
    cseb3_block1 = generate(c2f, ws2, "CSEB III First Half", seed+6, sync_sem3, room_prefix='C2', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=3,combined_sync=combined_sync_cse_sem3,semester_half=1, basket_used_global=basket_used_global, faculty_tt=faculty_tt, full_sem_sync=full_sem_sync_sem3, faculty_busy_global=faculty_busy_global[1])
    cseb3_block2 = generate(c2s, ws2, "CSEB III Second Half", seed+7, sync_sem3, room_prefix='C2', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=3,combined_sync=combined_sync_cse_sem3,semester_half=2, basket_used_global=basket_used_global, faculty_tt=faculty_tt, full_sem_sync=full_sem_sync_sem3, faculty_busy_global=faculty_busy_global[2])
    unscheduled += collect_unscheduled(c2f, cseb3_block1, "CSEB III First Half", year_tag=3, elective_sync=sync_sem3)
    unscheduled += collect_unscheduled(c2s, cseb3_block2, "CSEB III Second Half", year_tag=3, elective_sync=sync_sem3)
    add_csv_legend_block(ws2, coursesCSEB_III, "CSEB III - First Half", half=1, room_map_key="CSEB III First Half")
    add_csv_legend_block(ws2, coursesCSEB_III, "CSEB III - Second Half", half=2, room_map_key="CSEB III Second Half")

    
    combined_iii_courses = (csea3_block1 or []) + (csea3_block2 or []) + (cseb3_block1 or []) + (cseb3_block2 or [])
    reset_color_palette()
    merge_and_color(ws2, combined_iii_courses)

    # --- DSAI-III ---
    ws4 = wb.create_sheet("DSAI-III Timetable")
    d1f, d1s = split(coursesDSAI_III)
    dsa_block1 = generate(d1f, ws4, "DSAI-III First Half", seed+10, sync_sem3, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=3,combined_sync=combined_sync_de_sem3,semester_half=1, basket_used_global=basket_used_global, faculty_tt=faculty_tt, full_sem_sync=full_sem_sync_sem3, faculty_busy_global=faculty_busy_global[1])
    dsa_block2 = generate(d1s, ws4, "DSAI-III Second Half", seed+11, sync_sem3, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=3,combined_sync=combined_sync_de_sem3,semester_half=2, basket_used_global=basket_used_global, faculty_tt=faculty_tt, full_sem_sync=full_sem_sync_sem3, faculty_busy_global=faculty_busy_global[2])
    unscheduled += collect_unscheduled(d1f, dsa_block1, "DSAI-III First Half", year_tag=3, elective_sync=sync_sem3)
    unscheduled += collect_unscheduled(d1s, dsa_block2, "DSAI-III Second Half", year_tag=3, elective_sync=sync_sem3)
    add_csv_legend_block(ws4, coursesDSAI_III, "DSAI III - First Half", half=1, room_map_key="DSAI-III First Half")
    add_csv_legend_block(ws4, coursesDSAI_III, "DSAI III - Second Half", half=2, room_map_key="DSAI-III Second Half")

    combined_dsa_courses = (dsa_block1 or []) + (dsa_block2 or [])
    merge_and_color(ws4, combined_dsa_courses)

    # --- ECE-III ---
    ws5 = wb.create_sheet("ECE-III Timetable")
    e1f, e1s = split(coursesECE_III)
    ece_block1 = generate(e1f, ws5, "ECE-III First Half", seed+12, sync_sem3, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=3,combined_sync=combined_sync_de_sem3,semester_half=1, basket_used_global=basket_used_global, faculty_tt=faculty_tt, full_sem_sync=full_sem_sync_sem3, faculty_busy_global=faculty_busy_global[1])
    ece_block2 = generate(e1s, ws5, "ECE-III Second Half", seed+13, sync_sem3, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=3,combined_sync=combined_sync_de_sem3,semester_half=2, basket_used_global=basket_used_global, faculty_tt=faculty_tt, full_sem_sync=full_sem_sync_sem3, faculty_busy_global=faculty_busy_global[2])
    unscheduled += collect_unscheduled(e1f, ece_block1, "ECE-III First Half", year_tag=3, elective_sync=sync_sem3)
    unscheduled += collect_unscheduled(e1s, ece_block2, "ECE-III Second Half", year_tag=3, elective_sync=sync_sem3)
    add_csv_legend_block(ws5, coursesECE_III, "ECE III - First Half", half=1, room_map_key="ECE-III First Half")
    add_csv_legend_block(ws5, coursesECE_III, "ECE III - Second Half", half=2, room_map_key="ECE-III Second Half")

    combined_ece_courses = (ece_block1 or []) + (ece_block2 or [])
    merge_and_color(ws5, combined_ece_courses)

    # --- CSE-V ---
    ws3 = wb.create_sheet("CSE-V Timetable")
    c5af, c5as = split(coursesCSEA_V)
    c5bf, c5bs = split(coursesCSEB_V)
    c5a_block1 = generate(c5af, ws3, "CSEA V First Half", seed+8, sync_sem5_cse, room_prefix='C3', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=5,combined_sync=combined_sync_cse_sem5,semester_half=1, basket_used_global=basket_used_global, faculty_tt=faculty_tt, full_sem_sync=full_sem_sync_sem5, faculty_busy_global=faculty_busy_global[1])
    c5a_block2 = generate(c5as, ws3, "CSEA V Second Half", seed+9, sync_sem5_cse, room_prefix='C3', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=5,combined_sync=combined_sync_cse_sem5,semester_half=2, basket_used_global=basket_used_global, faculty_tt=faculty_tt, full_sem_sync=full_sem_sync_sem5, faculty_busy_global=faculty_busy_global[2])
    unscheduled += collect_unscheduled(c5af, c5a_block1, "CSEA V First Half", year_tag=5, elective_sync=sync_sem5_cse)
    unscheduled += collect_unscheduled(c5as, c5a_block2, "CSEA V Second Half", year_tag=5, elective_sync=sync_sem5_cse)
    add_csv_legend_block(ws3, coursesCSEA_V, "CSEA V - First Half", half=1, room_map_key="CSEA V First Half")
    add_csv_legend_block(ws3, coursesCSEA_V, "CSEA V - Second Half", half=2, room_map_key="CSEA V Second Half")

    c5b_block1 = generate(c5bf, ws3, "CSEB V First Half", seed+10, sync_sem5_cse, room_prefix='C3', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=5,combined_sync=combined_sync_cse_sem5,semester_half=1, basket_used_global=basket_used_global, faculty_tt=faculty_tt, full_sem_sync=full_sem_sync_sem5, faculty_busy_global=faculty_busy_global[1])
    c5b_block2 = generate(c5bs, ws3, "CSEB V Second Half", seed+11, sync_sem5_cse, room_prefix='C3', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=5,combined_sync=combined_sync_cse_sem5,semester_half=2, basket_used_global=basket_used_global, faculty_tt=faculty_tt, full_sem_sync=full_sem_sync_sem5, faculty_busy_global=faculty_busy_global[2])
    unscheduled += collect_unscheduled(c5bf, c5b_block1, "CSEB V First Half", year_tag=5, elective_sync=sync_sem5_cse)
    unscheduled += collect_unscheduled(c5bs, c5b_block2, "CSEB V Second Half", year_tag=5, elective_sync=sync_sem5_cse)
    add_csv_legend_block(ws3, coursesCSEB_V, "CSEB V - First Half", half=1, room_map_key="CSEB V First Half")
    add_csv_legend_block(ws3, coursesCSEB_V, "CSEB V - Second Half", half=2, room_map_key="CSEB V Second Half")

    combined_v_courses = (c5a_block1 or []) + (c5a_block2 or []) + (c5b_block1 or []) + (c5b_block2 or [])
    reset_color_palette()
    merge_and_color(ws3, combined_v_courses)

    ELECTIVE_SYNC_BY_YEAR[5] = sync_sem5_de

    # --- DSAI-V ---
    ws8 = wb.create_sheet("DSAI-V Timetable")
    d5f_v, d5s_v = split(coursesDSAI_V)
    dsai5_block1 = generate(d5f_v, ws8, "DSAI-V First Half", seed+18, sync_sem5_de, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=5,combined_sync=combined_sync_de_sem5,semester_half=1, basket_used_global=basket_used_global, faculty_tt=faculty_tt, full_sem_sync=full_sem_sync_sem5, faculty_busy_global=faculty_busy_global[1])
    dsai5_block2 = generate(d5s_v, ws8, "DSAI-V Second Half", seed+19, sync_sem5_de, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=5,combined_sync=combined_sync_de_sem5,semester_half=2, basket_used_global=basket_used_global, faculty_tt=faculty_tt, full_sem_sync=full_sem_sync_sem5, faculty_busy_global=faculty_busy_global[2])
    unscheduled += collect_unscheduled(d5f_v, dsai5_block1, "DSAI-V First Half", year_tag=5, elective_sync=sync_sem5_de)
    unscheduled += collect_unscheduled(d5s_v, dsai5_block2, "DSAI-V Second Half", year_tag=5, elective_sync=sync_sem5_de)
    add_csv_legend_block(ws8, coursesDSAI_V, "DSAI V - First Half", half=1, room_map_key="DSAI-V First Half")
    add_csv_legend_block(ws8, coursesDSAI_V, "DSAI V - Second Half", half=2, room_map_key="DSAI-V Second Half")

    combined_dsai5_courses = (dsai5_block1 or []) + (dsai5_block2 or [])
    merge_and_color(ws8, combined_dsai5_courses)

    # --- ECE-V ---
    ws10 = wb.create_sheet("ECE-V Timetable")
    e5f_v, e5s_v = split(coursesECE_V)
    ece5_block1 = generate(e5f_v, ws10, "ECE-V First Half", seed+22, sync_sem5_de, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=5,combined_sync=combined_sync_de_sem5,semester_half=1, basket_used_global=basket_used_global, faculty_tt=faculty_tt, full_sem_sync=full_sem_sync_sem5, faculty_busy_global=faculty_busy_global[1])
    ece5_block2 = generate(e5s_v, ws10, "ECE-V Second Half", seed+23, sync_sem5_de, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=5,combined_sync=combined_sync_de_sem5,semester_half=2, basket_used_global=basket_used_global, faculty_tt=faculty_tt, full_sem_sync=full_sem_sync_sem5, faculty_busy_global=faculty_busy_global[2])
    unscheduled += collect_unscheduled(e5f_v, ece5_block1, "ECE-V First Half", year_tag=5, elective_sync=sync_sem5_de)
    unscheduled += collect_unscheduled(e5s_v, ece5_block2, "ECE-V Second Half", year_tag=5, elective_sync=sync_sem5_de)
    add_csv_legend_block(ws10, coursesECE_V, "ECE V - First Half", half=1, room_map_key="ECE-V First Half")
    add_csv_legend_block(ws10, coursesECE_V, "ECE V - Second Half", half=2, room_map_key="ECE-V Second Half")

    combined_ece5_courses = (ece5_block1 or []) + (ece5_block2 or [])
    merge_and_color(ws10, combined_ece5_courses)
    # --- Common 7th Sem ---
    ws6 = wb.create_sheet("COMMON 7TH-SEM Timetable")
    s7f, s7s = split(coursesVII)
    s7_block1 = generate(s7f, ws6, "COMMON 7TH-SEM First Half", seed+14, sync_sem7, room_prefix='C3', elective_room_map=elective_room_map, room_busy_global=global_room_busy, year_tag=7, semester_half=1, basket_used_global=basket_used_global, faculty_tt=faculty_tt, full_sem_sync=full_sem_sync_sem7, faculty_busy_global=faculty_busy_global[1])
    s7_block2 = generate(s7s, ws6, "COMMON 7TH-SEM Second Half", seed+15, sync_sem7, room_prefix='C3', elective_room_map=elective_room_map, room_busy_global=global_room_busy, year_tag=7, semester_half=2, basket_used_global=basket_used_global, faculty_tt=faculty_tt, full_sem_sync=full_sem_sync_sem7, faculty_busy_global=faculty_busy_global[2])
    unscheduled += collect_unscheduled(s7f, s7_block1, "7TH SEM First Half", year_tag=7, elective_sync=sync_sem7)
    unscheduled += collect_unscheduled(s7s, s7_block2, "7TH SEM Second Half", year_tag=7, elective_sync=sync_sem7)
    add_csv_legend_block(ws6, coursesVII, "7TH SEM - First Half", half=1, room_map_key="COMMON 7TH-SEM First Half")
    add_csv_legend_block(ws6, coursesVII, "7TH SEM - Second Half", half=2, room_map_key="COMMON 7TH-SEM Second Half")
    combined_7_courses = (s7_block1 or []) + (s7_block2 or [])
    merge_and_color(ws6, combined_7_courses)

    name = f"Final_Timetable.xlsx"
    course_index = build_course_index()
    course_faculty_map = build_course_faculty_map()
    moved = repair_faculty_clashes(wb, faculty_tt, course_faculty_map)
    if moved:
        print(f"Clash repair moved {moved} entries to 17:30-18:30")
    wb.save(name)
    write_faculty_workbook(faculty_tt.get(1, {}), "Faculty_Timetable_First_Half.xlsx", course_index=course_index)
    write_faculty_workbook(faculty_tt.get(2, {}), "Faculty_Timetable_Second_Half.xlsx", course_index=course_index)

    # Export unscheduled courses report
    if unscheduled:
        df_uns = pd.DataFrame(unscheduled)
        df_uns = df_uns.drop_duplicates(subset=[
            "Department","Semester","Section","Course_Code","Course_Title",
            "Faculty","L-T-P-S-C","Elective","ElectiveBasket","Semester_Half"
        ])
        df_uns.to_excel("Unscheduled_Courses.xlsx", index=False)
    else:
        pd.DataFrame(columns=[
            "Group","Department","Semester","Section","Course_Code","Course_Title",
            "Faculty","L-T-P-S-C","Elective","ElectiveBasket","Semester_Half"
        ]).to_excel("Unscheduled_Courses.xlsx", index=False)
    print("Evenly balanced timetable saved in", name)


